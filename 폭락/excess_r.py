import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 파일 경로 설정 ────────────────────────────────────────────
BASE_DIR = r"C:\Users\intern6\trst_dev\truston_quant_dev\폭락"

FILE_0303        = rf"{BASE_DIR}\data_2949_20260303.csv"
FILE_0304        = rf"{BASE_DIR}\data_2913_20260304.csv"
FILE_0305        = rf"{BASE_DIR}\data_4814_20260306.csv"
FILE_KOSPI       = rf"{BASE_DIR}\kospi 200.csv"

FILE_KOSPI_SEC   = rf"{BASE_DIR}\코스피_업종_0303.csv"
FILE_KOSDAQ_SEC  = rf"{BASE_DIR}\코스닥_업종_0303.csv"

OUTPUT_0303          = rf"{BASE_DIR}\excess_return_20260303.csv"
OUTPUT_0304          = rf"{BASE_DIR}\excess_return_20260304.csv"
OUTPUT_0305          = rf"{BASE_DIR}\excess_return_20260305.csv"
OUTPUT_KOSPI_ALPHA   = rf"{BASE_DIR}\kospi_alpha_종목.csv"
OUTPUT_SECTOR_0303   = rf"{BASE_DIR}\섹터별_알파종목수_20260303.csv"
OUTPUT_SECTOR_0304   = rf"{BASE_DIR}\섹터별_알파종목수_20260304.csv"
OUTPUT_SECTOR_0305   = rf"{BASE_DIR}\섹터별_알파종목수_20260305.csv"
OUTPUT_SECTOR_EXCEL  = rf"{BASE_DIR}\업종별_알파종목.xlsx"
OUTPUT_SECTOR_COMPARE = rf"{BASE_DIR}\섹터별_알파종목수_비교.xlsx"
# ────────────────────────────────────────────────────────────────

def read_csv(path):
    try:
        return pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="euc-kr")

# 데이터 로드
df_0303  = read_csv(FILE_0303)
df_0304  = read_csv(FILE_0304)
df_0305  = read_csv(FILE_0305)
df_kospi = read_csv(FILE_KOSPI)

# 업종 데이터 (3일꺼 재사용 + 케이뱅크 수동 추가)
sector = pd.concat([read_csv(FILE_KOSPI_SEC), read_csv(FILE_KOSDAQ_SEC)], ignore_index=True)
kbank  = pd.DataFrame([{"종목코드": "279570", "시장구분": "KOSPI", "업종명": "은행"}])
sector = pd.concat([sector, kbank], ignore_index=True)

# 종목코드 타입 통일
for df in [df_0303, df_0304, df_0305, sector]:
    df["종목코드"] = df["종목코드"].astype(str).str.zfill(6)

# KOSPI 200 등락률 추출
kospi = df_kospi[["일자", "등락률"]].copy()
kospi["일자"] = kospi["일자"].str.replace("/", "-")
kospi = kospi.set_index("일자")["등락률"]

kospi_0303 = kospi.get("2026-03-03")
kospi_0304 = kospi.get("2026-03-04")
kospi_0305 = kospi.get("2026-03-05")

print(f"KOSPI200 등락률 → 03/03: {kospi_0303}%  /  03/04: {kospi_0304}%  /  03/05: {kospi_0305}%")

# ── 초과수익률 계산 + 업종/시장구분 병합 + 업종 없는 종목 제거 ──
def calc_excess(df, date_str, kospi_ret):
    df = df[["종목코드", "종목명", "등락률"]].copy()
    sector_cols = sector[["종목코드", "시장구분", "업종명"]].drop_duplicates("종목코드")
    df = df.merge(sector_cols, on="종목코드", how="inner")
    df["날짜"]           = date_str
    df["KOSPI200_등락률"] = round(kospi_ret, 2)
    df["초과수익률"]       = round(df["등락률"] - kospi_ret, 2)
    df = df.sort_values("초과수익률", ascending=False).reset_index(drop=True)
    df.insert(0, "순위", range(1, len(df) + 1))
    return df[["순위", "날짜", "종목코드", "종목명", "시장구분", "업종명", "등락률", "KOSPI200_등락률", "초과수익률"]]

df_0303_result = calc_excess(df_0303, "2026-03-03", kospi_0303)
df_0304_result = calc_excess(df_0304, "2026-03-04", kospi_0304)
df_0305_result = calc_excess(df_0305, "2026-03-05", kospi_0305)

# ── 전체 CSV 저장 ─────────────────────────────────────────────
df_0303_result.to_csv(OUTPUT_0303, index=False, encoding="utf-8-sig")
df_0304_result.to_csv(OUTPUT_0304, index=False, encoding="utf-8-sig")
df_0305_result.to_csv(OUTPUT_0305, index=False, encoding="utf-8-sig")

print(f"\n저장 완료:")
print(f"  {OUTPUT_0303}  ({len(df_0303_result)}종목)")
print(f"  {OUTPUT_0304}  ({len(df_0304_result)}종목)")
print(f"  {OUTPUT_0305}  ({len(df_0305_result)}종목)")

# ── KOSPI 알파 종목 추출 ──────────────────────────────────────
def get_kospi_alpha(df):
    out = df[(df["시장구분"] == "KOSPI") & (df["초과수익률"] > 0)].copy()
    out = out.sort_values("초과수익률", ascending=False).reset_index(drop=True)
    out["순위"] = range(1, len(out) + 1)
    return out

kospi_alpha_0303 = get_kospi_alpha(df_0303_result)
kospi_alpha_0304 = get_kospi_alpha(df_0304_result)
kospi_alpha_0305 = get_kospi_alpha(df_0305_result)

df_kospi_alpha = pd.concat([kospi_alpha_0303, kospi_alpha_0304, kospi_alpha_0305], ignore_index=True)
df_kospi_alpha = df_kospi_alpha[["순위", "날짜", "종목코드", "종목명", "업종명", "등락률", "KOSPI200_등락률", "초과수익률"]]
df_kospi_alpha.to_csv(OUTPUT_KOSPI_ALPHA, index=False, encoding="utf-8-sig")

print(f"  {OUTPUT_KOSPI_ALPHA}")
print(f"    03/03 KOSPI 알파: {len(kospi_alpha_0303)}개  /  03/04: {len(kospi_alpha_0304)}개  /  03/05: {len(kospi_alpha_0305)}개")

# ── 섹터별 알파 종목 수 집계 ──────────────────────────────────
def calc_sector_summary(df, date_str, kospi_ret):
    alpha = df[df["초과수익률"] > 0]
    pivot = alpha.groupby(["업종명", "시장구분"]).size().unstack(fill_value=0)
    for col in ["KOSDAQ", "KOSPI"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot[["KOSDAQ", "KOSPI"]]
    pivot["합계"] = pivot["KOSDAQ"] + pivot["KOSPI"]
    pivot = pivot.sort_values("합계", ascending=False).reset_index()
    pivot.insert(0, "날짜", date_str)
    pivot.insert(1, "KOSPI200_등락률", round(kospi_ret, 2))
    return pivot

sector_0303_summary = calc_sector_summary(df_0303_result, "2026-03-03", kospi_0303)
sector_0304_summary = calc_sector_summary(df_0304_result, "2026-03-04", kospi_0304)
sector_0305_summary = calc_sector_summary(df_0305_result, "2026-03-05", kospi_0305)

sector_0303_summary.to_csv(OUTPUT_SECTOR_0303, index=False, encoding="utf-8-sig")
sector_0304_summary.to_csv(OUTPUT_SECTOR_0304, index=False, encoding="utf-8-sig")
sector_0305_summary.to_csv(OUTPUT_SECTOR_0305, index=False, encoding="utf-8-sig")

print(f"  {OUTPUT_SECTOR_0303}")
print(f"  {OUTPUT_SECTOR_0304}")
print(f"  {OUTPUT_SECTOR_0305}")

# ── 업종별 알파 종목 엑셀 (3일치 통합) ───────────────────────
def write_sector_excel(date_results, output_path):
    COLS   = ["종목코드", "종목명", "시장구분", "등락률", "초과수익률"]
    HDR_BG = "1F3864"; HDR_FT = "FFFFFF"
    SEC_BG = "2E75B6"; SEC_FT = "FFFFFF"
    POS_FT = "1E6823"; NEG_FT = "C00000"
    THIN   = Side(style="thin", color="CCCCCC")
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    all_sectors = sorted(set().union(*[
        set(df[df["초과수익률"] > 0]["업종명"].unique()) for _, df, _ in date_results
    ]))

    for sector_name in all_sectors:
        ws = wb.create_sheet(title=sector_name[:31])
        for col, width in zip("ABCDE", [12, 18, 10, 10, 12]):
            ws.column_dimensions[col].width = width

        row = 1
        for date_str, df, kospi_ret in date_results:
            alpha_s = df[(df["업종명"] == sector_name) & (df["초과수익률"] > 0)].sort_values("초과수익률", ascending=False).reset_index(drop=True)

            # 날짜 헤더
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=1,
                value=f"{date_str}  |  {sector_name}  |  KOSPI200: {kospi_ret}%  |  알파 종목 {len(alpha_s)}개")
            cell.font      = Font(name="Arial", bold=True, size=10, color=HDR_FT)
            cell.fill      = PatternFill("solid", fgColor=HDR_BG)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1

            # 컬럼 헤더
            for c, h in enumerate(COLS, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font      = Font(name="Arial", bold=True, size=9, color=SEC_FT)
                cell.fill      = PatternFill("solid", fgColor=SEC_BG)
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            row += 1

            # 데이터
            if alpha_s.empty:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                ws.cell(row=row, column=1, value="알파 종목 없음").font = Font(name="Arial", italic=True, color="999999")
                row += 1
            else:
                for _, r in alpha_s.iterrows():
                    for c, val in enumerate([r["종목코드"], r["종목명"], r["시장구분"], r["등락률"], r["초과수익률"]], 1):
                        cell = ws.cell(row=row, column=c, value=val)
                        cell.font = Font(name="Arial", size=9,
                            color=POS_FT if (c in (4, 5) and val > 0) else (NEG_FT if (c in (4, 5) and val < 0) else "000000"))
                        cell.border    = border
                        cell.alignment = Alignment(horizontal="left" if c == 2 else "center")
                    row += 1
            row += 1  # 날짜 간 공백

        ws.freeze_panes = "A3"

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── 공통 알파 시트 추가 (합산 OR / 교집합 AND) ───────────────
    def _build_sector_lookup(date_results, days):
        day_sector_codes = {}
        day_sector_df    = {}
        for date_str, df, _ in date_results:
            if date_str not in days:
                continue
            alpha = df[df["초과수익률"] > 0]
            day_sector_codes[date_str] = {}
            day_sector_df[date_str]    = {}
            for sec, grp in alpha.groupby("업종명"):
                day_sector_codes[date_str][sec] = set(grp["종목코드"])
                day_sector_df[date_str][sec]    = grp
        return day_sector_codes, day_sector_df

    def get_union_alpha(date_results, days):
        """지정 날짜 중 하나라도 알파가 난 종목 (합산 OR) — 평균 등락률/초과수익률 표시"""
        day_sector_codes, day_sector_df = _build_sector_lookup(date_results, days)
        if not day_sector_codes:
            return {}
        all_sectors = sorted(set().union(*[set(v.keys()) for v in day_sector_codes.values()]))
        result = {}
        for sec in all_sectors:
            union_codes = set()
            for d in days:
                union_codes |= day_sector_codes.get(d, {}).get(sec, set())
            if not union_codes:
                result[sec] = pd.DataFrame()
                continue
            frames = [day_sector_df[d][sec][day_sector_df[d][sec]["종목코드"].isin(union_codes)]
                      for d in days if sec in day_sector_df.get(d, {})]
            combined = pd.concat(frames).groupby("종목코드").agg(
                종목명=("종목명", "first"),
                시장구분=("시장구분", "first"),
                등락률=("등락률", "mean"),
                초과수익률=("초과수익률", "mean"),
            ).reset_index()
            combined["등락률"]    = combined["등락률"].round(2)
            combined["초과수익률"] = combined["초과수익률"].round(2)
            result[sec] = combined.sort_values("초과수익률", ascending=False).reset_index(drop=True)
        return result

    def get_intersection_alpha(date_results, days):
        """지정 날짜 전부에서 알파가 난 종목 (교집합 AND) — 기준일(첫째 날) 수치 표시"""
        day_sector_codes, day_sector_df = _build_sector_lookup(date_results, days)
        if len(day_sector_codes) < len(days):
            return {}
        all_sectors = sorted(set().union(*[set(v.keys()) for v in day_sector_codes.values()]))
        result = {}
        for sec in all_sectors:
            sets = [day_sector_codes.get(d, {}).get(sec, set()) for d in days]
            inter_codes = sets[0]
            for s in sets[1:]:
                inter_codes &= s
            if not inter_codes:
                result[sec] = pd.DataFrame()
                continue
            base_df = day_sector_df[days[0]].get(sec, pd.DataFrame())
            result[sec] = base_df[base_df["종목코드"].isin(inter_codes)].sort_values("초과수익률", ascending=False).reset_index(drop=True)
        return result

    def write_alpha_sheet(wb, sector_dict, sheet_title, hdr_bg, sec_bg, sub_label):
        HDR_FT = "FFFFFF"; POS_FT = "1E6823"; NEG_FT = "C00000"
        COLS   = ["종목코드", "종목명", "시장구분", "등락률", "초과수익률"]
        THIN   = Side(style="thin", color="CCCCCC")
        border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        ws = wb.create_sheet(title=sheet_title)
        for col, width in zip("ABCDE", [12, 18, 10, 10, 12]):
            ws.column_dimensions[col].width = width

        row = 1
        for sec in sorted(sector_dict.keys()):
            df_sec = sector_dict[sec]
            cnt = len(df_sec)

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=1, value=f"{sec}  |  {sub_label}  |  {cnt}개")
            cell.font      = Font(name="Arial", bold=True, size=10, color=HDR_FT)
            cell.fill      = PatternFill("solid", fgColor=hdr_bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1

            for c, h in enumerate(COLS, 1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font      = Font(name="Arial", bold=True, size=9, color=HDR_FT)
                cell.fill      = PatternFill("solid", fgColor=sec_bg)
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            row += 1

            if df_sec.empty:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                cell = ws.cell(row=row, column=1, value="해당 종목 없음")
                cell.font = Font(name="Arial", italic=True, color="999999", size=9)
                cell.alignment = Alignment(horizontal="center")
                row += 1
            else:
                for _, r in df_sec.iterrows():
                    for c, val in enumerate([r["종목코드"], r["종목명"], r["시장구분"], r["등락률"], r["초과수익률"]], 1):
                        cell = ws.cell(row=row, column=c, value=val)
                        cell.font = Font(name="Arial", size=9,
                            color=POS_FT if (c in (4, 5) and isinstance(val, (int, float)) and val > 0)
                                  else (NEG_FT if (c in (4, 5) and isinstance(val, (int, float)) and val < 0)
                                  else "000000"))
                        cell.border    = border
                        cell.alignment = Alignment(horizontal="left" if c == 2 else "center")
                    row += 1
            row += 1

        ws.freeze_panes = "A3"
        total = sum(len(v) for v in sector_dict.values())
        print(f"  [{sheet_title}] 총 {total}개")

    # 합산(OR): 3·4일 중 하나라도 알파 → 평균 수치
    write_alpha_sheet(wb,
        get_union_alpha(date_results, ["2026-03-03", "2026-03-04"]),
        "📌 3·4일 합산 알파", "7B3F00", "C87941", "3·4일 합산(OR) 알파")
    # 합산(OR): 3·4·5일 중 하나라도 알파 → 평균 수치
    write_alpha_sheet(wb,
        get_union_alpha(date_results, ["2026-03-03", "2026-03-04", "2026-03-05"]),
        "📌 3·4·5일 합산 알파", "4A235A", "8E44AD", "3·4·5일 합산(OR) 알파")
    # 교집합(AND): 3·4일 모두 알파 → 3/3 기준 수치
    write_alpha_sheet(wb,
        get_intersection_alpha(date_results, ["2026-03-03", "2026-03-04"]),
        "📌 3·4일 교집합 알파", "1A3A5C", "2E75B6", "3·4일 교집합(AND) 알파")
    # 교집합(AND): 3·4·5일 모두 알파 → 3/3 기준 수치
    write_alpha_sheet(wb,
        get_intersection_alpha(date_results, ["2026-03-03", "2026-03-04", "2026-03-05"]),
        "📌 3·4·5일 교집합 알파", "1A4731", "2E7D52", "3·4·5일 교집합(AND) 알파")

    wb.save(output_path)

date_results = [
    ("2026-03-03", df_0303_result, kospi_0303),
    ("2026-03-04", df_0304_result, kospi_0304),
    ("2026-03-05", df_0305_result, kospi_0305),
]
write_sector_excel(date_results, OUTPUT_SECTOR_EXCEL)
print(f"  {OUTPUT_SECTOR_EXCEL}  (3일치 통합)")

# ── 미리보기 ─────────────────────────────────────────────────
for date, df in [("2026-03-03", df_0303_result), ("2026-03-04", df_0304_result), ("2026-03-05", df_0305_result)]:
    print(f"\n{'='*55}")
    print(f"[{date}]  KOSPI200: {df['KOSPI200_등락률'].iloc[0]}%  /  총 {len(df)}종목")
    print(f"\n▲ 초과수익 상위 10종목")
    print(df[["순위", "종목명", "시장구분", "업종명", "등락률", "초과수익률"]].head(10).to_string(index=False))
    print(f"\n▼ 초과수익 하위 10종목")
    print(df[["순위", "종목명", "시장구분", "업종명", "등락률", "초과수익률"]].tail(10).to_string(index=False))

print(f"\n{'='*55}")
print("[섹터별 알파 종목 수]")
for date, df in [("2026-03-03", sector_0303_summary), ("2026-03-04", sector_0304_summary), ("2026-03-05", sector_0305_summary)]:
    print(f"\n{date}")
    print(df[["업종명", "KOSDAQ", "KOSPI", "합계"]].to_string(index=False))

print(f"\n{'='*55}")
print("모든 파일 저장 완료!")

# ── 섹터별 알파 종목 수 날짜 비교 엑셀 ───────────────────────
def write_sector_compare_excel(summaries, output_path):
    THIN = Side(style="thin", color="CCCCCC")
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    DATE_META = [
        (f"2026-03-03  (KOSPI200: {kospi_0303}%)", "0303", "1F3864", "D6E4F0"),
        (f"2026-03-04  (KOSPI200: {kospi_0304}%)", "0304", "7B1818", "FAD7D7"),
        (f"2026-03-05  (KOSPI200: +{kospi_0305}%)", "0305", "1A4731", "D6F0E4"),
    ]

    all_sectors = sorted(set().union(*[set(s["업종명"]) for s in summaries]))
    merged = pd.DataFrame({"업종명": all_sectors})
    for (_, key, _, _), s in zip(DATE_META, summaries):
        s = s.rename(columns={"KOSDAQ": f"KQ_{key}", "KOSPI": f"KS_{key}", "합계": f"합계_{key}"})
        merged = merged.merge(s[["업종명", f"KQ_{key}", f"KS_{key}", f"합계_{key}"]], on="업종명", how="left").fillna(0)
    merged["_sort"] = sum(merged[f"합계_{k}"] for _, k, _, _ in DATE_META)
    merged = merged.sort_values("_sort", ascending=False).drop("_sort", axis=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "섹터별 알파종목수 비교"

    ws.merge_cells("A1:A2")
    cell = ws["A1"]
    cell.value = "업종명"
    cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="333333")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border

    col = 2
    for label, key, hdr_color, sub_color in DATE_META:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        for c_off, sub in enumerate(["KOSDAQ", "KOSPI", "합계"], 0):
            cell2 = ws.cell(row=2, column=col+c_off, value=sub)
            cell2.font      = Font(name="Arial", bold=True, size=9)
            cell2.fill      = PatternFill("solid", fgColor=sub_color)
            cell2.alignment = Alignment(horizontal="center")
            cell2.border    = border
        col += 3

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16
    ws.column_dimensions["A"].width = 18
    for c in range(2, 12):
        ws.column_dimensions[chr(64+c)].width = 9

    for r_idx, row in merged.iterrows():
        excel_row = r_idx + 3
        cell = ws.cell(row=excel_row, column=1, value=row["업종명"])
        cell.font      = Font(name="Arial", bold=True, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill      = PatternFill("solid", fgColor="F7F7F7" if r_idx % 2 == 0 else "FFFFFF")
        cell.border    = border

        col = 2
        for _, key, _, sub_color in DATE_META:
            for c_off, col_key in enumerate([f"KQ_{key}", f"KS_{key}", f"합계_{key}"], 0):
                val = int(row[col_key])
                cell = ws.cell(row=excel_row, column=col+c_off, value=val if val > 0 else "-")
                cell.font      = Font(name="Arial", size=9,
                    bold=(c_off == 2 and val > 0),
                    color="1F3864" if (c_off == 2 and val > 0) else ("1E6823" if val > 0 else "999999"))
                cell.alignment = Alignment(horizontal="center")
                cell.fill      = PatternFill("solid", fgColor=sub_color if c_off < 2 else "FFFFFF")
                cell.border    = border
            col += 3

    total_row = len(merged) + 3
    cell = ws.cell(row=total_row, column=1, value="전체 합계")
    cell.font = Font(name="Arial", bold=True, size=9)
    cell.fill = PatternFill("solid", fgColor="E8E8E8")
    cell.alignment = Alignment(horizontal="left")
    cell.border = border

    col = 2
    for _, key, _, _ in DATE_META:
        for c_off, col_key in enumerate([f"KQ_{key}", f"KS_{key}", f"합계_{key}"], 0):
            total = int(merged[col_key].sum())
            cell = ws.cell(row=total_row, column=col+c_off, value=total)
            cell.font      = Font(name="Arial", bold=True, size=9, color="1F3864")
            cell.alignment = Alignment(horizontal="center")
            cell.fill      = PatternFill("solid", fgColor="E8E8E8")
            cell.border    = border
        col += 3

    ws.freeze_panes = "A3"
    wb.save(output_path)

write_sector_compare_excel(
    [sector_0303_summary, sector_0304_summary, sector_0305_summary],
    OUTPUT_SECTOR_COMPARE
)
print(f"  {OUTPUT_SECTOR_COMPARE}  (3일치 비교)")