"""
ETF 분석 스크립트 — 폭락장 테마/시총/거래대금 분석 + HTML + Excel
사용법: python etf_report.py
"""
import pandas as pd, numpy as np, os, warnings
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ═══ 경로 설정 (여기만 수정) ═══
DATA_DIR = r"C:\Users\intern6\trst_dev\truston_quant_dev\폭락\data"
RESULT_DIR = r"C:\Users\intern6\trst_dev\truston_quant_dev\폭락\results"
os.makedirs(RESULT_DIR, exist_ok=True)

ETF_FILES = {
    "03": os.path.join(DATA_DIR, "etf_3.3.csv"),
    "04": os.path.join(DATA_DIR, "etf_3.4.csv"),
    "05": os.path.join(DATA_DIR, "etf_3.5.csv"),
    "06": os.path.join(DATA_DIR, "etf_3.6.csv"),
    "09": os.path.join(DATA_DIR, "etf_3.9.csv"),
    "10": os.path.join(DATA_DIR, "etf_3.10.csv"),
}

# ═══ 유틸 ═══
def read_csv(path):
    for enc in ["utf-8","euc-kr","cp949"]:
        try: return pd.read_csv(path, encoding=enc)
        except: continue
    raise FileNotFoundError(f"Cannot read: {path}")

# ═══ 1. 데이터 로드 ═══
print("▶ ETF 데이터 로드...")
dfs = {}
for d, f in ETF_FILES.items():
    df = read_csv(f); df["종목코드"] = df["종목코드"].astype(str).str.zfill(6); dfs[d] = df
    print(f"  3/{d}: {len(df)}종목")

base = dfs["03"][["종목코드","종목명","시가총액","기초지수_지수명"]].copy()
for d in ["03","04","05","06","09","10"]:
    base = base.merge(dfs[d][["종목코드","등락률","거래대금"]].rename(
        columns={"등락률":f"rt_{d}","거래대금":f"val_{d}"}), on="종목코드", how="left")

# ═══ 2. 필터 ═══
print("▶ 해외/채권/인버스/레버리지 필터링...")
exclude_kw = [
    "미국","나스닥","S&P","NYSE","다우","필라델피아","Russell",
    "중국","차이나","CSI","항셍","H주","홍콩","심천",
    "일본","니케이","TOPIX","닛케이",
    "유럽","독일","DAX","EURO","프랑스","영국",
    "인도","베트남","대만","브라질","멕시코","인니","글로벌","선진국","이머징",
    "달러","엔화","위안","유로","통화","환노출","환헤지",
    "WTI","원유","천연가스","금선물","은선물","구리","팔라듐","니켈","금현물",
    "비트코인","이더리움","블록체인","크립토",
    "테슬라","엔비디아","애플","마이크로소프트","아마존","구글","메타","샤오미",
    "FANG","빅테크","필리핀","태국",
    "채권","금리","CD","국고","통안","머니마켓","단기금융","회사채","국채",
    "소버린","크레딧","하이일드","만기자동","롱숏","만기매칭","특수채",
    "국공채","공사채","종합채","중장기","단기채","금융채","은행채",
    "싱가포르","일라이릴리","WGBI","리츠부동산","FTSE","MSCI",
    "인버스","곰","숏","Bear","레버리지","2X","불","Bull",
    "TDF","TRF","전단채","골드선물","농산물","콩선물","팔란티어",
    "버크셔","토탈월드","단기자금","커버드콜",
]

def is_ex(name):
    for kw in exclude_kw:
        if kw.upper() in str(name).upper(): return True
    return False

base["제외"] = base["종목명"].apply(is_ex)
etf = base[~base["제외"]].copy()

# ═══ 3. 계산 ═══
print("▶ 계산...")
etf["시총_억"] = etf["시가총액"]/1e8
etf["평균거래대금_억"] = etf[[f"val_{d}" for d in ["03","04","05","06","09","10"]]].mean(axis=1)/1e8
etf["하락avg"] = etf[["rt_03","rt_04","rt_09"]].mean(axis=1)
etf["반등avg"] = etf[["rt_05","rt_10"]].mean(axis=1)
etf["6일합"] = etf[["rt_03","rt_04","rt_05","rt_06","rt_09","rt_10"]].sum(axis=1)

# 테마 분류
def theme(name):
    name=str(name)
    if any(k in name for k in ["방산","방위","국방","K-방산","우주항공","K방산"]): return "방산/우주"
    if any(k in name for k in ["2차전지","배터리","양극재","리튬","친환경차"]): return "2차전지"
    if any(k in name for k in ["반도체","AI","HBM","핵심공정","후공정","전공정","소부장","온디바이스"]): return "반도체/AI"
    if any(k in name for k in ["조선","해운","해양"]): return "조선/해운"
    if any(k in name for k in ["원자력","원전","핵","우라늄","SMR"]): return "원자력"
    if any(k in name for k in ["바이오","헬스","제약","의료","게놈"]): return "바이오"
    if any(k in name for k in ["자동차","EV","전기차","모빌리티","현대차"]): return "자동차"
    if any(k in name for k in ["에너지","신재생","태양","풍력","수소"]): return "에너지"
    if any(k in name for k in ["은행","금융","보험","증권"]): return "금융"
    if any(k in name for k in ["배당","밸류","가치","고배당","주주환원"]): return "배당/가치"
    if any(k in name for k in ["200","코스피","KOSPI","KRX","코리아"]): return "시장지수"
    if any(k in name for k in ["코스닥","150","KOSDAQ"]): return "코스닥"
    if any(k in name for k in ["게임","미디어","엔터","콘텐츠","KPOP"]): return "엔터/게임"
    if any(k in name for k in ["로봇","자율주행"]): return "로봇"
    if any(k in name for k in ["철강","소재","화학","포스코"]): return "소재"
    if any(k in name for k in ["건설","인프라","SOC","네트워크"]): return "건설/인프라"
    if any(k in name for k in ["통신","5G","IT"]): return "통신"
    if any(k in name for k in ["소비","유통","식품"]): return "소비"
    if any(k in name for k in ["삼성","LG","SK","한화","롯데","현대"]): return "그룹주"
    if any(k in name for k in ["Top","TOP","대형","5대"]): return "대형주"
    return "기타"

etf["테마"] = etf["종목명"].apply(theme)
etf["시총3"] = pd.qcut(etf["시총_억"],3,labels=["소형","중형","대형"])
etf["거래대금3"] = pd.qcut(etf["평균거래대금_억"],3,labels=["저","중","고"])
etf["그룹"] = etf["시총3"].astype(str)+"/"+etf["거래대금3"].astype(str)

print(f"  국내 ETF: {len(etf)}종목, 테마: {etf['테마'].nunique()}개")

# ═══ 4. Excel ═══
print("▶ Excel 생성...")
THIN=Side(style="thin",color="CCCCCC"); bdr=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
GRP_COLORS={"소형/저":"E8D5B7","소형/중":"E8D5B7","소형/고":"E8D5B7",
    "중형/저":"D4E6F1","중형/중":"D4E6F1","중형/고":"D4E6F1",
    "대형/저":"D5F5E3","대형/중":"D5F5E3","대형/고":"D5F5E3"}

wb=Workbook()

# Sheet1: 전체 (반등avg 순)
ws=wb.active; ws.title="ETF_전체(반등순)"
headers=["#","시총그룹","거래대금","테마","종목명","시총(억)","거래대금(억)",
         "3/3","3/4","3/5","3/6","3/9","3/10","하락avg","반등avg","6일합"]
widths=[4,8,8,12,32,8,8,7,7,7,7,7,7,8,8,8]
for c,(h,w) in enumerate(zip(headers,widths),1):
    cell=ws.cell(row=1,column=c,value=h)
    cell.font=Font("Arial",bold=True,size=9,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor="333333"); cell.border=bdr
    cell.alignment=Alignment(horizontal="center"); ws.column_dimensions[get_column_letter(c)].width=w

for i,(_,row) in enumerate(etf.sort_values("반등avg",ascending=False).iterrows(),1):
    r=i+1; color=GRP_COLORS.get(row["그룹"],"FFFFFF")
    vals=[i,str(row["시총3"]),str(row["거래대금3"]),row["테마"],row["종목명"],
          round(row["시총_억"],0),round(row["평균거래대금_억"],1),
          row.get("rt_03"),row.get("rt_04"),row.get("rt_05"),row.get("rt_06"),row.get("rt_09"),row.get("rt_10"),
          round(row["하락avg"],2),round(row["반등avg"],2),round(row["6일합"],2)]
    for c,v in enumerate(vals,1):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font("Arial",size=9,color="CC0000" if isinstance(v,(int,float)) and v is not None and c>=8 and v<0 else ("008800" if isinstance(v,(int,float)) and v is not None and c>=8 and v>0 else "000000"))
        cell.fill=PatternFill("solid",fgColor=color); cell.border=bdr
        cell.alignment=Alignment(horizontal="center" if c!=5 else "left")
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{len(etf)+1}"

# Sheet2: 테마별 (반등avg 순)
ws2=wb.create_sheet("테마별(반등순)")
th=etf.groupby("테마").agg(n=("종목코드","count"),r=("반등avg","mean"),b1=("rt_05","mean"),b2=("rt_10","mean"),
    wr1=("rt_05",lambda x:(x>0).mean()*100),d=("하락avg","mean"),t=("6일합","mean"),
    e03=("rt_03","mean"),e04=("rt_04","mean"),e05=("rt_05","mean"),e06=("rt_06","mean"),e09=("rt_09","mean"),e10=("rt_10","mean")).round(2)
th=th.sort_values("r",ascending=False)
h2=["#","테마","ETF수","반등avg","3/5","3/10","승률3/5","하락avg","6일합","3/3","3/4","3/5 ","3/6","3/9","3/10 "]
for c,(h,w) in enumerate(zip(h2,[4,12,6,8,7,7,8,8,8,7,7,7,7,7,7]),1):
    cell=ws2.cell(row=1,column=c,value=h); cell.font=Font("Arial",bold=True,size=9,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor="333333"); cell.border=bdr; cell.alignment=Alignment(horizontal="center")
    ws2.column_dimensions[get_column_letter(c)].width=w
for i,(tm,row) in enumerate(th.iterrows(),1):
    vals=[i,tm,row["n"],row["r"],row["b1"],row["b2"],row["wr1"],row["d"],row["t"],
          row["e03"],row["e04"],row["e05"],row["e06"],row["e09"],row["e10"]]
    for c,v in enumerate(vals,1):
        cell=ws2.cell(row=i+1,column=c,value=v); cell.font=Font("Arial",size=9)
        cell.border=bdr; cell.alignment=Alignment(horizontal="center" if c!=2 else "left")

# Sheet3: 시총×거래대금 (반등avg 순)
ws3=wb.create_sheet("시총x거래대금(반등순)")
rows9=[]
for g in ["소형","중형","대형"]:
    for l in ["저","중","고"]:
        gs=etf[(etf["시총3"]==g)&(etf["거래대금3"]==l)]
        if len(gs)<3: continue
        rows9.append({"g":g,"l":l,"n":len(gs),"r":gs["반등avg"].mean(),"b1":gs["rt_05"].mean(),
            "b2":gs["rt_10"].mean(),"wr1":(gs["rt_05"]>0).mean()*100,"d":gs["하락avg"].mean(),"t":gs["6일합"].mean()})
rows9.sort(key=lambda x:x["r"],reverse=True)
h3=["#","시총","거래대금","종목수","반등avg","3/5","3/10","승률3/5","하락avg","6일합"]
for c,(h,w) in enumerate(zip(h3,[4,8,8,6,8,7,7,8,8,8]),1):
    cell=ws3.cell(row=1,column=c,value=h); cell.font=Font("Arial",bold=True,size=9,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor="333333"); cell.border=bdr; cell.alignment=Alignment(horizontal="center")
    ws3.column_dimensions[get_column_letter(c)].width=w
for i,row in enumerate(rows9,1):
    color=GRP_COLORS.get(f"{row['g']}/{row['l']}","FFFFFF")
    vals=[i,row["g"],row["l"],row["n"],round(row["r"],2),round(row["b1"],1),round(row["b2"],1),
          round(row["wr1"],1),round(row["d"],2),round(row["t"],2)]
    for c,v in enumerate(vals,1):
        cell=ws3.cell(row=i+1,column=c,value=v); cell.font=Font("Arial",size=9)
        cell.fill=PatternFill("solid",fgColor=color); cell.border=bdr; cell.alignment=Alignment(horizontal="center")

if "Sheet" in wb.sheetnames: del wb["Sheet"]
excel_path=os.path.join(RESULT_DIR,"ETF_시총_거래대금_분석.xlsx")
wb.save(excel_path)
print(f"  → {excel_path}")

# ═══ 5. HTML ═══
print("▶ HTML 생성...")
def cls(v): return "pos" if v>0.01 else "neg" if v<-0.01 else ""

CSS="""*{margin:0;padding:0;box-sizing:border-box}body{background:#070d1b;color:#e2e8f0;font-family:'Noto Sans KR',sans-serif;line-height:1.6}
.S{min-height:100vh;padding:44px 56px;display:flex;flex-direction:column;justify-content:center;border-bottom:1px solid #1e293b}
.S:nth-child(even){background:#0b1120}h1{font-size:34px;font-weight:900;letter-spacing:-1.5px;margin-bottom:10px}h1 span{color:#3b82f6}
h2{font-size:22px;font-weight:700;margin-bottom:14px}h2 b{display:inline-block;background:#3b82f6;color:#fff;width:28px;height:28px;border-radius:50%;text-align:center;line-height:28px;font-size:13px;margin-right:7px}
.sub{color:#64748b;font-size:12px;margin-bottom:14px}
.hl{background:linear-gradient(120deg,rgba(239,68,68,0.1),rgba(59,130,246,0.1));border-left:4px solid #ef4444;padding:10px 16px;border-radius:0 8px 8px 0;margin:10px 0;font-size:12px}.hl em{color:#f8fafc;font-style:normal;font-weight:700}
.hl.b{border-left-color:#3b82f6}.hl.g{border-left-color:#10b981}
.sr{display:flex;gap:8px;flex-wrap:wrap;margin:8px 0}.st{background:#1e293b;border-radius:8px;padding:10px 12px;flex:1;min-width:90px;border:1px solid #334155}
.st .lb{font-size:9px;color:#64748b}.st .vl{font-size:18px;font-weight:700;margin:2px 0}.st .ds{font-size:9px;color:#94a3b8}
.pos{color:#10b981}.neg{color:#ef4444}
table{width:100%;border-collapse:collapse;margin:6px 0;font-size:11px}th{background:#1e293b;color:#94a3b8;padding:6px 7px;text-align:center;font-weight:600;border-bottom:2px solid #334155;white-space:nowrap}
td{padding:5px 7px;text-align:center;border-bottom:1px solid #1e293b;white-space:nowrap}tr:hover{background:rgba(59,130,246,0.04)}.tl{text-align:left}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px}.cd{background:rgba(30,41,59,0.5);border-radius:10px;padding:14px;border:1px solid #1e293b}
.ct{font-size:12px;font-weight:700;margin-bottom:6px}"""

H=f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>ETF 폭락 분석</title><link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
<style>{CSS}@media print{{.S{{page-break-after:always;min-height:auto;padding:20px}}}}</style></head><body>"""

# P1
H+=f"""<div class="S" style="justify-content:center;align-items:center;text-align:center">
<h1>🇮🇷 미·이란 전쟁 위기 <span>국내 ETF</span> 분석</h1>
<p class="sub" style="font-size:14px">2026.03.03~03.10 · {len(etf)} ETF (해외/채권/인버스/레버리지 제외)</p></div>"""

# P2 테마별
def theme_tbl(sort_col):
    ts=th.sort_values(sort_col,ascending=False)
    t=f'<table><tr><th>#</th><th class="tl">테마</th><th>ETF</th><th>반등avg</th><th>3/5</th><th>3/10</th><th>승률</th><th>하락avg</th><th>6일합</th></tr>'
    for i,(tm,r) in enumerate(ts.iterrows(),1):
        t+=f'<tr><td>{i}</td><td class="tl">{tm}</td><td>{r["n"]:.0f}</td><td class="pos"><b>+{r["r"]:.2f}%</b></td><td class="pos">+{r["b1"]:.1f}%</td><td class="pos">+{r["b2"]:.1f}%</td><td>{r["wr1"]:.0f}%</td><td class="neg">{r["d"]:+.2f}%</td><td class="{cls(r["t"])}">{r["t"]:+.2f}%</td></tr>'
    t+='</table>'; return t

H+=f"""<div class="S"><h2><b>1</b>테마별 — 6일합 순 vs 반등avg 순</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">6일합 순</div>{theme_tbl("t")}</div>
<div class="cd"><div class="ct" style="color:#10b981">반등avg 순</div>{theme_tbl("r")}</div>
</div></div>"""

# P3 시총×거래대금
def grp_tbl(sk):
    rs=sorted(rows9,key=lambda x:x[sk],reverse=True)
    t='<table><tr><th>#</th><th>시총</th><th>거래대금</th><th>n</th><th>반등avg</th><th>3/5</th><th>3/10</th><th>승률</th><th>하락avg</th><th>6일합</th></tr>'
    for i,r in enumerate(rs,1):
        t+=f'<tr><td>{i}</td><td>{r["g"]}</td><td>{r["l"]}</td><td>{r["n"]}</td><td class="pos"><b>+{r["r"]:.2f}%</b></td><td class="pos">+{r["b1"]:.1f}%</td><td class="pos">+{r["b2"]:.1f}%</td><td>{r["wr1"]:.0f}%</td><td class="neg">{r["d"]:+.2f}%</td><td class="{cls(r["t"])}">{r["t"]:+.2f}%</td></tr>'
    t+='</table>'; return t

H+=f"""<div class="S"><h2><b>2</b>시총 × 거래대금 — 6일합 순 vs 반등avg 순</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">6일합 순</div>{grp_tbl("t")}</div>
<div class="cd"><div class="ct" style="color:#10b981">반등avg 순</div>{grp_tbl("r")}</div>
</div></div>"""

# P4 t-test
def vol_cmp(metric):
    t=f'<table><tr><th>시총</th><th>저</th><th>중</th><th>고</th><th>고-저</th><th>p</th><th>유의?</th></tr>'
    for g in ["소형","중형","대형"]:
        lo=etf[(etf["시총3"]==g)&(etf["거래대금3"]=="저")][metric]
        hi=etf[(etf["시총3"]==g)&(etf["거래대금3"]=="고")][metric]
        mi=etf[(etf["시총3"]==g)&(etf["거래대금3"]=="중")][metric]
        try: _,pv=stats.ttest_ind(lo,hi); diff=hi.mean()-lo.mean()
        except: pv=1; diff=0
        sig="✅" if pv<0.05 else "❌"
        t+=f'<tr><td>{g}</td><td>{lo.mean():+.2f}%</td><td>{mi.mean():+.2f}%</td><td>{hi.mean():+.2f}%</td><td class="{cls(diff)}">{diff:+.2f}%p</td><td>{pv:.3f}</td><td>{sig}</td></tr>'
    t+='</table>'; return t

H+=f"""<div class="S"><h2><b>3</b>같은 시총 내 거래대금별 차이 — t-test</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#10b981">반등avg 기준</div>{vol_cmp("반등avg")}</div>
<div class="cd"><div class="ct" style="color:#3b82f6">6일합 기준</div>{vol_cmp("6일합")}</div>
</div></div>"""

# P5 TOP 20
top_r=etf[etf["시총_억"]>=300].nlargest(20,"반등avg")
top_t=etf[etf["시총_억"]>=300].nlargest(20,"6일합")
def top_tbl(df,col):
    t=f'<table><tr><th>#</th><th class="tl">ETF</th><th>테마</th><th>시총</th><th>반등avg</th><th>하락avg</th><th>6일합</th></tr>'
    for i,(_,r) in enumerate(df.iterrows(),1):
        t+=f'<tr><td>{i}</td><td class="tl">{r["종목명"]}</td><td>{r["테마"]}</td><td>{r["시총_億"]:,.0f}</td><td class="pos">+{r["반등avg"]:.1f}%</td><td class="neg">{r["하락avg"]:+.1f}%</td><td class="{cls(r["6일합"])}">{r["6일합"]:+.1f}%</td></tr>'
    t+='</table>'; return t
# 키 보정
top_r=top_r.copy(); top_r["시총_億"]=top_r["시총_억"]
top_t=top_t.copy(); top_t["시총_億"]=top_t["시총_억"]

H+=f"""<div class="S"><h2><b>4</b>ETF TOP 20 (시총 300억+)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#10b981">반등avg 순</div>{top_tbl(top_r,"반등avg")}</div>
<div class="cd"><div class="ct" style="color:#3b82f6">6일합 순</div>{top_tbl(top_t,"6일합")}</div>
</div></div>"""

# P6 결론
H+="""<div class="S" style="align-items:center;text-align:center">
<h2 style="font-size:26px">ETF 분석 결론</h2>
<div style="max-width:660px;text-align:left;margin:14px auto">
<div style="display:flex;gap:10px;margin:10px 0;padding:10px;background:#1e293b;border-radius:8px">
<div style="font-size:18px">🎯</div><div><b style="font-size:12px">알파는 테마</b><br>
<span style="color:#94a3b8;font-size:11px">시총/거래대금이 아니라 방산·반도체·원자력 같은 테마가 수익률 결정.</span></div></div>
<div style="display:flex;gap:10px;margin:10px 0;padding:10px;background:#1e293b;border-radius:8px">
<div style="font-size:18px">📊</div><div><b style="font-size:12px">반등매매 = 반도체/AI ETF</b><br>
<span style="color:#94a3b8;font-size:11px">반등avg 최고. 하락에서도 지수 수준이라 리스크 적음.</span></div></div>
<div style="display:flex;gap:10px;margin:10px 0;padding:10px;background:#1e293b;border-radius:8px">
<div style="font-size:18px">🛡️</div><div><b style="font-size:12px">6일 전체 = 방산 ETF</b><br>
<span style="color:#94a3b8;font-size:11px">하락일에도 방어, 반등일에도 상승. 유일한 양수 테마.</span></div></div>
<div style="display:flex;gap:10px;margin:10px 0;padding:10px;background:#1e293b;border-radius:8px">
<div style="font-size:18px">⚠️</div><div><b style="font-size:12px">거래대금 = 독립 변수 아님</b><br>
<span style="color:#94a3b8;font-size:11px">t-test 전 그룹 유의하지 않음. 개별종목과 동일 결론.</span></div></div>
</div></div>"""

H+="</body></html>"
html_path=os.path.join(RESULT_DIR,"ETF_report.html")
with open(html_path,"w",encoding="utf-8") as f: f.write(H)
print(f"  → {html_path}")

print(f"\n✅ 완료!")
print(f"  Excel: {excel_path}")
print(f"  HTML:  {html_path}")