"""
종목 분석 스크립트 — 폭락장 시총/거래대금 분석 + HTML + Excel
사용법: python stock_report.py
"""
import pandas as pd, numpy as np, os, warnings
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ═══ 경로 설정 (여기만 수정) ═══
DATA_DIR = r"C:\Users\intern6\trst_dev\truston_quant_dev\폭락\data"
RESULT_DIR = r"C:\Users\intern6\trst_dev\truston_quant_dev\폭락\results"
os.makedirs(RESULT_DIR, exist_ok=True)

# 날짜→파일 매핑 (파일명 수정 시 여기만 변경)
FILES = {
    "2026-03-03": os.path.join(DATA_DIR, "data_2949_20260303.csv"),
    "2026-03-04": os.path.join(DATA_DIR, "data_2913_20260304.csv"),
    "2026-03-05": os.path.join(DATA_DIR, "data_4814_20260305.csv"),  # 파일명 오류, 실제 3/5 데이터
    "2026-03-06": os.path.join(DATA_DIR, "data_4018_20260306.csv"),
    "2026-03-09": os.path.join(DATA_DIR, "data_3913_20260309.csv"),
    "2026-03-10": os.path.join(DATA_DIR, "data_5408_20260310.csv"),
}
KOSPI_FILE = os.path.join(DATA_DIR, "kospi.csv")
KOSDAQ_FILE = os.path.join(DATA_DIR, "kosdaq.csv")
SECTOR_FILES = {
    "KOSPI": os.path.join(DATA_DIR, "코스피_업종_0303.csv"),
    "KOSDAQ": os.path.join(DATA_DIR, "코스닥_업종_0303.csv"),
}

DATES = sorted(FILES.keys())

# ═══ 유틸 ═══
def read_csv(path):
    for enc in ["utf-8","euc-kr","cp949"]:
        try: return pd.read_csv(path, encoding=enc)
        except: continue
    raise FileNotFoundError(f"Cannot read: {path}")

# ═══ 1. 데이터 로드 ═══
print("▶ 데이터 로드...")

# 지수
ki = read_csv(KOSPI_FILE); ki.columns = range(len(ki.columns))
ki["일자"] = ki[0].str.replace("/","-"); ki["등락률"] = ki[3].astype(float)
kospi_ret = ki.set_index("일자")["등락률"].to_dict()

kq = read_csv(KOSDAQ_FILE); kq["일자"] = kq["일자"].str.replace("/","-")
kosdaq_ret = kq.set_index("일자")["등락률"].to_dict()

# 업종
sector = pd.concat([read_csv(SECTOR_FILES["KOSPI"])[["종목코드","시장구분","업종명","시가총액"]],
                     read_csv(SECTOR_FILES["KOSDAQ"])[["종목코드","시장구분","업종명","시가총액"]],
                     pd.DataFrame([{"종목코드":"279570","시장구분":"KOSPI","업종명":"은행","시가총액":0}])])
sector["종목코드"] = sector["종목코드"].astype(str).str.zfill(6)
sector = sector.drop_duplicates("종목코드")
sector["시가총액"] = pd.to_numeric(sector["시가총액"], errors="coerce")

# 종목별 일별 데이터
all_results = []
for date_str, fpath in sorted(FILES.items()):
    df = read_csv(fpath)
    df["종목코드"] = df["종목코드"].astype(str).str.zfill(6)
    df = df.merge(sector[["종목코드","시장구분","업종명","시가총액"]], on="종목코드", how="inner")
    df = df[df["시장구분"].isin(["KOSPI","KOSDAQ"])].copy()
    df["벤치마크"] = df["시장구분"].map({"KOSPI":kospi_ret.get(date_str,0),"KOSDAQ":kosdaq_ret.get(date_str,0)})
    df["초과수익률"] = round(df["등락률"] - df["벤치마크"], 2)
    df["날짜"] = date_str
    all_results.append(df)
    print(f"  {date_str}: {len(df)}종목")
all_df = pd.concat(all_results, ignore_index=True)

# ═══ 2. Master 구축 ═══
print("▶ Master 구축...")
piv_val = all_df.pivot_table(index="종목코드", columns="날짜", values="거래대금")
piv_rt = all_df.pivot_table(index="종목코드", columns="날짜", values="등락률")

m = all_df.drop_duplicates("종목코드")[["종목코드","종목명","시장구분","업종명"]].copy()
m = m.merge(sector[["종목코드","시가총액"]], on="종목코드", how="left")
m["시총_억"] = m["시가총액"]/1e8
m["평균거래대금_억"] = piv_val.mean(axis=1).values/1e8

for d in DATES:
    k = d[8:]
    m = m.merge(all_df[all_df["날짜"]==d][["종목코드","초과수익률","등락률"]].rename(
        columns={"초과수익률":f"ex_{k}","등락률":f"rt_{k}"}), on="종목코드", how="left")

m["하락avg"] = m[["rt_03","rt_04","rt_09"]].mean(axis=1)
m["반등avg"] = m[["rt_05","rt_10"]].mean(axis=1)
m["6일합"] = m[["rt_03","rt_04","rt_05","rt_06","rt_09","rt_10"]].sum(axis=1)

for mk in ["KOSPI","KOSDAQ"]:
    mask = m["시장구분"]==mk
    m.loc[mask,"Q5"] = pd.qcut(m.loc[mask,"시총_억"],5,labels=["Q1_최소","Q2_소형","Q3_중형","Q4_대형","Q5_최대"])
    m.loc[mask,"거래대금3"] = pd.qcut(m.loc[mask,"평균거래대금_억"],3,labels=["저","중","고"])

m["그룹"] = m["Q5"].astype(str).str.split("_").str[1] + "/" + m["거래대금3"].astype(str)

B = {"KOSPI":{"d":np.mean([kospi_ret[d] for d in ["2026-03-03","2026-03-04","2026-03-09"]]),
              "r":np.mean([kospi_ret[d] for d in ["2026-03-05","2026-03-10"]]),
              "t":sum(kospi_ret[d] for d in [f"2026-03-{x}" for x in ["03","04","05","06","09","10"]])},
     "KOSDAQ":{"d":np.mean([kosdaq_ret[d] for d in ["2026-03-03","2026-03-04","2026-03-09"]]),
               "r":np.mean([kosdaq_ret[d] for d in ["2026-03-05","2026-03-10"]]),
               "t":sum(kosdaq_ret[d] for d in [f"2026-03-{x}" for x in ["03","04","05","06","09","10"]])}}

print(f"  KOSPI: {len(m[m['시장구분']=='KOSPI'])}종목, KOSDAQ: {len(m[m['시장구분']=='KOSDAQ'])}종목")

# ═══ 3. Excel 생성 ═══
print("▶ Excel 생성...")

THIN = Side(style="thin",color="CCCCCC"); bdr = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
GRP_COLORS = {"최소/저":"E8D5B7","최소/중":"E8D5B7","최소/고":"E8D5B7",
    "소형/저":"D4E6F1","소형/중":"D4E6F1","소형/고":"D4E6F1",
    "중형/저":"D5F5E3","중형/중":"D5F5E3","중형/고":"D5F5E3",
    "대형/저":"FADBD8","대형/중":"FADBD8","대형/고":"FADBD8",
    "최대/저":"D7BDE2","최대/중":"D7BDE2","최대/고":"D7BDE2"}

wb = Workbook()
for mk in ["KOSPI","KOSDAQ"]:
    sub = m[m["시장구분"]==mk].dropna(subset=["Q5","거래대금3"]).copy()
    sub = sub.sort_values(["Q5","거래대금3","반등avg"], ascending=[True,True,False])
    ws = wb.active if mk=="KOSPI" else wb.create_sheet()
    ws.title = mk
    headers = ["시총그룹","거래대금","종목코드","종목명","업종","시총(억)","거래대금(억)",
               "3/3","3/4","3/5","3/6","3/9","3/10","하락avg","반등avg","6일합"]
    widths = [10,8,10,16,14,10,10,8,8,8,8,8,8,8,8,8]
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(row=1,column=c,value=h)
        cell.font=Font("Arial",bold=True,size=9,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="333333"); cell.border=bdr
        cell.alignment=Alignment(horizontal="center"); ws.column_dimensions[get_column_letter(c)].width=w
    r=2
    for _,row in sub.iterrows():
        grp=row["그룹"]; color=GRP_COLORS.get(grp,"FFFFFF")
        vals=[grp.split("/")[0],grp.split("/")[1],row["종목코드"],row["종목명"],row["업종명"],
              round(row["시총_억"],0),round(row["평균거래대금_억"],1),
              row.get("rt_03"),row.get("rt_04"),row.get("rt_05"),row.get("rt_06"),row.get("rt_09"),row.get("rt_10"),
              round(row["하락avg"],2),round(row["반등avg"],2),round(row["6일합"],2)]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=r,column=c,value=v)
            cell.font=Font("Arial",size=9,color="CC0000" if isinstance(v,(int,float)) and v is not None and c>=8 and v<0 else ("008800" if isinstance(v,(int,float)) and v is not None and c>=8 and v>0 else "000000"))
            cell.fill=PatternFill("solid",fgColor=color); cell.border=bdr
            cell.alignment=Alignment(horizontal="center" if c>4 else "left")
        r+=1
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{r-1}"

    # 요약 시트
    ws2=wb.create_sheet(f"{mk}_요약")
    sh=["#","시총","거래대금","종목수","반등avg(%)","3/5(%)","3/10(%)","승률3/5(%)","하락avg(%)","6일합(%)"]
    for c,(h,w) in enumerate(zip(sh,[4,8,8,6,10,8,8,10,10,8]),1):
        cell=ws2.cell(row=1,column=c,value=h)
        cell.font=Font("Arial",bold=True,size=9,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="333333"); cell.border=bdr
        cell.alignment=Alignment(horizontal="center"); ws2.column_dimensions[get_column_letter(c)].width=w
    rows=[]
    for g in ["Q1_최소","Q2_소형","Q3_중형","Q4_대형","Q5_최대"]:
        for l in ["저","중","고"]:
            gs=sub[(sub["Q5"]==g)&(sub["거래대금3"]==l)]
            if len(gs)<3: continue
            rows.append({"g":g.split("_")[1],"l":l,"n":len(gs),"r":gs["반등avg"].mean(),
                "b1":gs["rt_05"].mean(),"b2":gs["rt_10"].mean(),"wr1":(gs["rt_05"]>0).mean()*100,
                "d":gs["하락avg"].mean(),"t":gs["6일합"].mean()})
    rows.sort(key=lambda x:x["r"],reverse=True)
    for i,row in enumerate(rows,1):
        color=GRP_COLORS.get(f"{row['g']}/{row['l']}","FFFFFF")
        vals=[i,row["g"],row["l"],row["n"],round(row["r"],2),round(row["b1"],1),round(row["b2"],1),
              round(row["wr1"],1),round(row["d"],2),round(row["t"],2)]
        for c,v in enumerate(vals,1):
            cell=ws2.cell(row=i+1,column=c,value=v)
            cell.font=Font("Arial",size=9); cell.fill=PatternFill("solid",fgColor=color)
            cell.border=bdr; cell.alignment=Alignment(horizontal="center")

if "Sheet" in wb.sheetnames: del wb["Sheet"]
excel_path = os.path.join(RESULT_DIR, "시총_거래대금_15그룹.xlsx")
wb.save(excel_path)
print(f"  → {excel_path}")

# ═══ 4. HTML 생성 ═══
print("▶ HTML 생성...")

def cls(v): return "pos" if v>0.01 else "neg" if v<-0.01 else ""

# 데이터 준비 함수들
def q5data(mk):
    s=m[m["시장구분"]==mk]; rows=[]
    for g in ["Q1_최소","Q2_소형","Q3_중형","Q4_대형","Q5_최대"]:
        gs=s[s["Q5"]==g]; rows.append({"g":g.split("_")[1],"n":len(gs),
            "d":round(gs["하락avg"].mean(),2),"r":round(gs["반등avg"].mean(),2),"t":round(gs["6일합"].mean(),2),
            "dv":round(gs["하락avg"].mean()-B[mk]["d"],2),"rv":round(gs["반등avg"].mean()-B[mk]["r"],2),
            "b1":round(gs["rt_05"].mean(),1),"b2":round(gs["rt_10"].mean(),1),
            "wr1":round((gs["rt_05"]>0).mean()*100,0),"wr2":round((gs["rt_10"]>0).mean()*100,0),
            "d04":round(gs["rt_04"].mean(),1),"d09":round(gs["rt_09"].mean(),1)})
    return rows

def q5l_data(mk):
    s=m[m["시장구분"]==mk].dropna(subset=["거래대금3"]); rows=[]
    for g in ["Q1_최소","Q2_소형","Q3_중형","Q4_대형","Q5_최대"]:
        for l in ["저","중","고"]:
            gs=s[(s["Q5"]==g)&(s["거래대금3"]==l)]
            if len(gs)<5: continue
            rows.append({"g":g.split("_")[1],"l":l,"n":len(gs),
                "d":round(gs["하락avg"].mean(),2),"r":round(gs["반등avg"].mean(),2),"t":round(gs["6일합"].mean(),2),
                "b1":round(gs["rt_05"].mean(),1),"wr1":round((gs["rt_05"]>0).mean()*100,0)})
    return rows

def bounce_stocks(mk,min_cap=0,min_val=0,n=15):
    s=m[(m["시장구분"]==mk)&(m["시총_억"]>=min_cap)&(m["평균거래대금_억"]>=min_val)].copy()
    s["bavg"]=(s["rt_05"]+s["rt_10"])/2
    return s.nlargest(n,"bavg")[["종목명","업종명","시총_억","평균거래대금_억","rt_05","rt_10","bavg"]].round(1).to_dict("records")

def vol_compare(mk):
    sub=m[m["시장구분"]==mk].dropna(subset=["거래대금3"]); t=''
    t+='<table style="font-size:11px"><tr><th>시총</th><th>저</th><th>중</th><th>고</th><th>고-저</th><th>p-value</th><th>유의?</th></tr>'
    for g in ["Q1_최소","Q2_소형","Q3_중형","Q4_대형","Q5_최대"]:
        gs=sub[sub["Q5"]==g]
        lo=gs[gs["거래대금3"]=="저"]["반등avg"]; hi=gs[gs["거래대금3"]=="고"]["반등avg"]; mi=gs[gs["거래대금3"]=="중"]["반등avg"]
        try: _,pv=stats.ttest_ind(lo,hi); diff=hi.mean()-lo.mean()
        except: pv=1; diff=0
        sig="✅" if pv<0.05 else "❌"; gl=g.split("_")[1]
        t+=f'<tr><td>{gl}</td><td>{lo.mean():+.2f}%</td><td>{mi.mean():+.2f}%</td><td>{hi.mean():+.2f}%</td><td class="{cls(diff)}">{diff:+.2f}%p</td><td>{pv:.3f}</td><td>{sig}</td></tr>'
    t+='</table>'; return t

ks5=q5data("KOSPI"); kq5=q5data("KOSDAQ")
ks5l=q5l_data("KOSPI"); kq5l=q5l_data("KOSDAQ")

# HTML 조립 (presentation_v4.html과 동일 구조)
# ... (이하 HTML 생성 코드는 위에서 만든 것과 동일하므로 파일에서 읽어옴)

# 간략화를 위해 핵심 테이블 생성 함수만
def tbl_q5(data,bk,sort_key="t"):
    s=sorted(data,key=lambda x:x[sort_key],reverse=True)
    bn=f'<span style="background:#1e293b;padding:4px 8px;border-radius:4px;font-size:10px;display:inline-block;margin:3px">지수 하락avg {bk["d"]:+.2f}% / 반등avg {bk["r"]:+.2f}% / 6일 {bk["t"]:+.2f}%</span>'
    t=bn+'<table><tr><th>#</th><th>그룹</th><th>n</th><th>하락avg</th><th>반등avg</th><th>6일합</th><th>vs지수하락</th><th>vs지수반등</th></tr>'
    for i,r in enumerate(s,1):
        t+=f'<tr><td>{i}</td><td>{r["g"]}</td><td>{r["n"]}</td><td class="neg">{r["d"]:+.2f}%</td><td class="pos">{r["r"]:+.2f}%</td><td class="{cls(r["t"])}"><b>{r["t"]:+.2f}%</b></td><td class="{cls(r["dv"])}">{r["dv"]:+.2f}%</td><td class="{cls(r["rv"])}">{r["rv"]:+.2f}%</td></tr>'
    t+='</table>'; return t

def tbl_q5l(data,sort_key="t"):
    s=sorted(data,key=lambda x:x[sort_key],reverse=True)
    t='<table style="font-size:11px"><tr><th>#</th><th>시총</th><th>거래대금</th><th>n</th><th>반등avg</th><th>3/5</th><th>3/10</th><th>승률</th><th>하락avg</th><th>6일합</th></tr>'
    for i,r in enumerate(s,1):
        t+=f'<tr><td>{i}</td><td>{r["g"]}</td><td>{r["l"]}</td><td>{r["n"]}</td><td class="pos"><b>+{r["r"]:.2f}%</b></td><td class="pos">+{r["b1"]:.1f}%</td><td class="neg">{r.get("b2",0):+.1f}%</td><td>{r["wr1"]:.0f}%</td><td class="neg">{r["d"]:+.2f}%</td><td class="{cls(r["t"])}">{r["t"]:+.2f}%</td></tr>'
    t+='</table>'; return t

def tbl_stocks(data):
    t='<table><tr><th>#</th><th class="tl">종목</th><th>업종</th><th>시총</th><th>거래대금</th><th>3/5</th><th>3/10</th><th>반등avg</th></tr>'
    for i,r in enumerate(data,1):
        t+=f'<tr><td>{i}</td><td class="tl">{r["종목명"]}</td><td>{r["업종명"]}</td><td>{r["시총_億"]:,.0f}</td><td>{r["평균거래대금_億"]:.0f}</td><td class="pos">+{r["rt_05"]:.1f}%</td><td class="pos">+{r["rt_10"]:.1f}%</td><td class="pos"><b>+{r["bavg"]:.1f}%</b></td></tr>'
    t+='</table>'; return t

def tbl_bounce(data):
    s=sorted(data,key=lambda x:(x["b1"]+x["b2"])/2,reverse=True)
    t='<table><tr><th>#</th><th>그룹</th><th>n</th><th>3/4하락</th><th>3/5반등</th><th>승률</th><th>3/9하락</th><th>3/10반등</th><th>승률</th><th>반등avg</th></tr>'
    for i,r in enumerate(s,1):
        a=(r["b1"]+r["b2"])/2
        t+=f'<tr><td>{i}</td><td>{r["g"]}</td><td>{r["n"]}</td><td class="neg">{r["d04"]:+.1f}%</td><td class="pos">+{r["b1"]:.1f}%</td><td>{r["wr1"]:.0f}%</td><td class="neg">{r["d09"]:+.1f}%</td><td class="pos">+{r["b2"]:.1f}%</td><td>{r["wr2"]:.0f}%</td><td class="pos"><b>+{a:.1f}%</b></td></tr>'
    t+='</table>'; return t

# HTML 전체 조립
CSS = """*{margin:0;padding:0;box-sizing:border-box}body{background:#070d1b;color:#e2e8f0;font-family:'Noto Sans KR',sans-serif;line-height:1.6}
.S{min-height:100vh;padding:44px 56px;display:flex;flex-direction:column;justify-content:center;border-bottom:1px solid #1e293b}
.S:nth-child(even){background:#0b1120}h1{font-size:36px;font-weight:900;letter-spacing:-1.5px;margin-bottom:10px}h1 span{color:#ef4444}
h2{font-size:24px;font-weight:700;margin-bottom:16px}h2 b{display:inline-block;background:#3b82f6;color:#fff;width:30px;height:30px;border-radius:50%;text-align:center;line-height:30px;font-size:14px;margin-right:8px}
.sub{color:#64748b;font-size:13px;margin-bottom:16px}
.hl{background:linear-gradient(120deg,rgba(239,68,68,0.1),rgba(59,130,246,0.1));border-left:4px solid #ef4444;padding:12px 18px;border-radius:0 8px 8px 0;margin:12px 0;font-size:13px}.hl em{color:#f8fafc;font-style:normal;font-weight:700}
.hl.b{border-left-color:#3b82f6}.hl.g{border-left-color:#10b981}
.sr{display:flex;gap:10px;flex-wrap:wrap;margin:10px 0}.st{background:#1e293b;border-radius:8px;padding:12px;flex:1;min-width:100px;border:1px solid #334155}
.st .lb{font-size:10px;color:#64748b}.st .vl{font-size:20px;font-weight:700;margin:2px 0}.st .ds{font-size:10px;color:#94a3b8}
.pos{color:#10b981}.neg{color:#ef4444}
table{width:100%;border-collapse:collapse;margin:8px 0;font-size:11.5px}th{background:#1e293b;color:#94a3b8;padding:7px 8px;text-align:center;font-weight:600;border-bottom:2px solid #334155;white-space:nowrap}
td{padding:6px 8px;text-align:center;border-bottom:1px solid #1e293b;white-space:nowrap}tr:hover{background:rgba(59,130,246,0.04)}.tl{text-align:left}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:16px}.cd{background:rgba(30,41,59,0.5);border-radius:10px;padding:16px;border:1px solid #1e293b}
.ct{font-size:13px;font-weight:700;margin-bottom:8px}"""

H = f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>종목 폭락 분석</title><link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
<style>{CSS}@media print{{.S{{page-break-after:always;min-height:auto;padding:20px}}}}</style></head><body>"""

# Page 1-11 (동일 구조)
idx_data = [("KOSPI",-7.24,-12.06,9.63,0.02,-5.96,5.35,-4.62,-14.0,14.1,3.43,-4.54,3.21)]
H += f"""<div class="S" style="justify-content:center;align-items:center;text-align:center">
<h1>🇮🇷 미·이란 전쟁 위기 <span>한국 증시</span> 분석</h1>
<p class="sub" style="font-size:15px">2026.03.03~03.10 · 6거래일 · {len(m)}종목</p></div>"""

H += f"""<div class="S"><h2><b>1</b>시장 개요</h2>
<div class="sr">
<div class="st"><div class="lb">3/3 폭락1</div><div class="vl neg">-7.24%</div><div class="ds">KQ -4.62%</div></div>
<div class="st"><div class="lb">3/4 폭락2</div><div class="vl neg">-12.06%</div><div class="ds">KQ -14.00%</div></div>
<div class="st"><div class="lb">3/5 반등1</div><div class="vl pos">+9.63%</div><div class="ds">KQ +14.10%</div></div>
<div class="st"><div class="lb">3/6 횡보</div><div class="vl" style="color:#94a3b8">+0.02%</div><div class="ds">KQ +3.43%</div></div>
<div class="st"><div class="lb">3/9 재폭락</div><div class="vl neg">-5.96%</div><div class="ds">KQ -4.54%</div></div>
<div class="st"><div class="lb">3/10 반등2</div><div class="vl pos">+5.35%</div><div class="ds">KQ +3.21%</div></div>
</div></div>"""

H += f"""<div class="S"><h2><b>2</b>시총별 등락률 (6일합 수익률 순)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI</div>{tbl_q5(ks5,B["KOSPI"],"t")}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ</div>{tbl_q5(kq5,B["KOSDAQ"],"t")}</div>
</div></div>"""

H += f"""<div class="S"><h2><b>3</b>시총 × 거래대금 교차 (6일합 순)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI</div>{tbl_q5l(ks5l,"t")}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ</div>{tbl_q5l(kq5l,"t")}</div>
</div></div>"""

H += f"""<div class="S"><h2><b>4</b>두 번의 반등 비교 (반등avg 순)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI</div>{tbl_bounce(ks5)}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ</div>{tbl_bounce(kq5)}</div>
</div></div>"""

ks_bt=bounce_stocks("KOSPI",500); kq_bt=bounce_stocks("KOSDAQ",500)
ks_bl=bounce_stocks("KOSPI",0,10); kq_bl=bounce_stocks("KOSDAQ",0,10)

# 종목 테이블용 키 보정
for lst in [ks_bt,kq_bt,ks_bl,kq_bl]:
    for r in lst:
        r["시총_億"]=r.pop("시총_억",0); r["평균거래대금_億"]=r.pop("평균거래대금_억",0)

H += f"""<div class="S"><h2><b>5</b>반등매매 TOP (시총 500억+)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI</div>{tbl_stocks(ks_bt)}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ</div>{tbl_stocks(kq_bt)}</div>
</div></div>"""

H += f"""<div class="S"><h2><b>6</b>반등매매 TOP (거래대금 10억+)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI</div>{tbl_stocks(ks_bl)}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ</div>{tbl_stocks(kq_bl)}</div>
</div></div>"""

# 반등avg 순 페이지
H += f"""<div class="S"><h2><b>7</b>시총별 — 반등일만 기준 (반등avg 순)</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI</div>{tbl_q5(ks5,B["KOSPI"],"r")}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ</div>{tbl_q5(kq5,B["KOSDAQ"],"r")}</div>
</div></div>"""

H += f"""<div class="S"><h2><b>8</b>시총 × 거래대금 — 반등avg 순 + t-test</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI (반등avg 순)</div>{tbl_q5l(ks5l,"r")}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ (반등avg 순)</div>{tbl_q5l(kq5l,"r")}</div>
</div>
<div class="g2" style="margin-top:12px">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI t-test</div>{vol_compare("KOSPI")}</div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ t-test</div>{vol_compare("KOSDAQ")}</div>
</div></div>"""

H += "</body></html>"

html_path = os.path.join(RESULT_DIR, "stock_report.html")
with open(html_path, "w", encoding="utf-8") as f:
    f.write(H)
print(f"  → {html_path}")

print(f"\n✅ 완료!")
print(f"  Excel: {excel_path}")
print(f"  HTML:  {html_path}")