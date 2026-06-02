"""
원본 CSV → 엑셀 + HTML 생성 (단일 스크립트)
실행: python 3.3_3.9_report.py
출력: results\final_conclusions.xlsx + 3.3_3.9.html
"""
import pandas as pd
import numpy as np
import os
import warnings
warnings.filterwarnings("ignore")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 ──
ROOT = r"C:\Users\intern6\trst_dev\truston_quant_dev\폭락"
DATA = rf"{ROOT}\data"
RESULTS = rf"{ROOT}\results"

EXCEL_OUT = rf"{RESULTS}\final_conclusions.xlsx"
HTML_OUT = rf"{ROOT}\3.3_3.9.html"

FILES = {
    "2026-03-03": rf"{DATA}\data_2949_20260303.csv",
    "2026-03-04": rf"{DATA}\data_2913_20260304.csv",
    "2026-03-05": rf"{DATA}\data_4814_20260305.csv",
    "2026-03-06": rf"{DATA}\data_4018_20260306.csv",
    "2026-03-09": rf"{DATA}\data_3913_20260309.csv",
}
KOSPI_SEC  = rf"{DATA}\코스피_업종_0303.csv"
KOSDAQ_SEC = rf"{DATA}\코스닥_업종_0303.csv"
KOSPI_IDX  = rf"{DATA}\kospi.csv"
KOSDAQ_IDX = rf"{DATA}\kosdaq.csv"

def read_csv(path):
    for enc in ["utf-8", "euc-kr", "cp949"]:
        try: return pd.read_csv(path, encoding=enc)
        except: continue

# ── Load ──
sector = pd.concat([
    read_csv(KOSPI_SEC)[["종목코드", "시장구분", "업종명", "시가총액"]],
    read_csv(KOSDAQ_SEC)[["종목코드", "시장구분", "업종명", "시가총액"]],
])
kbank = pd.DataFrame([{"종목코드": "279570", "시장구분": "KOSPI", "업종명": "은행", "시가총액": 0}])
sector = pd.concat([sector, kbank], ignore_index=True)
sector["종목코드"] = sector["종목코드"].astype(str).str.zfill(6)
sector = sector.drop_duplicates("종목코드")
sector["시가총액"] = pd.to_numeric(sector["시가총액"], errors="coerce")

kospi_idx = read_csv(KOSPI_IDX); kosdaq_idx = read_csv(KOSDAQ_IDX)
kospi_idx["일자"] = kospi_idx["일자"].str.replace("/", "-"); kosdaq_idx["일자"] = kosdaq_idx["일자"].str.replace("/", "-")
kospi_ret = kospi_idx.set_index("일자")["등락률"].to_dict()
kosdaq_ret = kosdaq_idx.set_index("일자")["등락률"].to_dict()

all_results = []
for date_str, fpath in sorted(FILES.items()):
    df = read_csv(fpath)
    df["종목코드"] = df["종목코드"].astype(str).str.zfill(6)
    df = df.merge(sector[["종목코드", "시장구분", "업종명", "시가총액"]], on="종목코드", how="inner")
    df = df[df["시장구분"].isin(["KOSPI", "KOSDAQ"])].copy()
    df["벤치마크_등락률"] = df["시장구분"].map({
        "KOSPI": kospi_ret.get(date_str, 0), "KOSDAQ": kosdaq_ret.get(date_str, 0)})
    df["초과수익률"] = round(df["등락률"] - df["벤치마크_등락률"], 2)
    df["날짜"] = date_str
    all_results.append(df)
    print(f"  {date_str}: {len(df)}종목")
all_df = pd.concat(all_results, ignore_index=True)

DATES = ["2026-03-03", "2026-03-04", "2026-03-05", "2026-03-06", "2026-03-09"]
piv_ex = all_df.pivot_table(index="종목코드", columns="날짜", values="초과수익률")
piv_rt = all_df.pivot_table(index="종목코드", columns="날짜", values="등락률")
piv_val = all_df.pivot_table(index="종목코드", columns="날짜", values="거래대금")

meta = all_df.drop_duplicates("종목코드")[["종목코드", "종목명", "시장구분", "업종명"]].copy()
meta = meta.merge(sector[["종목코드", "시가총액"]], on="종목코드", how="left")
meta["시총_억"] = meta["시가총액"] / 1e8
meta["평균거래대금_억"] = piv_val.mean(axis=1).values / 1e8
meta["절대변동성"] = piv_rt.std(axis=1).values

for d in DATES:
    k = d[8:]
    meta = meta.merge(all_df[all_df["날짜"]==d][["종목코드", "초과수익률", "등락률"]].rename(
        columns={"초과수익률": f"ex_{k}", "등락률": f"rt_{k}"}), on="종목코드", how="left")

meta["하락avg"] = meta[["ex_03", "ex_04", "ex_09"]].mean(axis=1)
meta["반등"] = meta["ex_05"]
meta["횡보"] = meta["ex_06"]
meta["동등스코어"] = ((meta["하락avg"] + meta["반등"]) / 2).round(2)
meta["반등1일"] = meta["rt_05"]
meta["반등2일"] = ((1 + meta["rt_05"]/100) * (1 + meta["rt_06"]/100) - 1) * 100

def qc(r):
    if r["하락avg"] > 0 and r["반등"] > 0: return "Q1"
    elif r["하락avg"] < 0 and r["반등"] > 0: return "Q2"
    elif r["하락avg"] < 0 and r["반등"] < 0: return "Q3"
    else: return "Q4"
meta["사분면"] = meta.apply(qc, axis=1)

for m in ["KOSPI", "KOSDAQ"]:
    mask = meta["시장구분"] == m
    meta.loc[mask, "Q5"] = pd.qcut(meta.loc[mask, "시총_억"], 5, labels=["1_최소", "2_소형", "3_중형", "4_대형", "5_최대"])
    meta.loc[mask, "유동성5"] = pd.qcut(meta.loc[mask, "평균거래대금_억"], 5, labels=["1_극소", "2_소", "3_중", "4_대", "5_극대"])

print(f"Master: {len(meta)} stocks loaded")

# ══════════════════════════════════════
# EXCEL
# ══════════════════════════════════════
THIN = Side(style="thin", color="334155")
bdr = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FONT = Font("Arial", bold=True, size=9, color="f8fafc")
H_FILL = PatternFill("solid", fgColor="1e293b")
POS = Font("Arial", size=9, color="10b981", bold=True)
NEG = Font("Arial", size=9, color="ef4444", bold=True)
NRM = Font("Arial", size=9, color="e2e8f0")
BG0 = PatternFill("solid", fgColor="0f172a")
BG1 = PatternFill("solid", fgColor="1a2332")

def hdr(ws, row, cols, widths=None):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def stitle(ws, row, title, ncols, color="3b82f6"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font("Arial", bold=True, size=11, color="f8fafc")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26
    return row + 1

def wrow(ws, row, vals, pn=None):
    pn = pn or []
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c)
        cell.value = round(v, 2) if isinstance(v, float) else v
        if c in pn and isinstance(v, (int, float)):
            cell.font = POS if v > 0 else NEG if v < 0 else NRM
        else:
            cell.font = NRM
        cell.fill = BG1 if row % 2 == 0 else BG0; cell.border = bdr
        cell.alignment = Alignment(horizontal="center" if c > 2 else "left")

wb = Workbook()

# ── Sheet 1: 결론 요약 ──
ws = wb.active; ws.title = "결론요약"
ws.sheet_properties.tabColor = "ef4444"
r = 1
r = stitle(ws, r, "미·이란 전쟁 위기 한국 증시 분석 — 핵심 결론 6가지", 8, "ef4444"); r += 1
conclusions = [
    ("결론 1", "소형주 방어력은 착시", "KOSDAQ 극소유동 하락일 양수 98.1% → 반등일 양수 0.0%. 거래 부재로 가격이 안 움직인 것."),
    ("결론 2", "시총별 동등비교하면 알파 없음", "KOSPI 전 그룹 순합 ±0.25% 이내. 하락일 방어 = 반등일 미달. 상관계수 -0.989."),
    ("결론 3", "반등매매는 대형주가 최적", "시총 클수록 반등 수익·승률 단조증가. KOSDAQ 최대: -14%→+13.2%, 승률99.5%."),
    ("결론 4", "두 폭락은 성격이 다름", "3/4 무차별 투매 vs 3/9 선별적 매도. 시장이 학습함."),
    ("결론 5", "에너지 테마 = 역방향 베팅", "하락3일 +30% 초과수익 → 반등일 -10~20% 되돌림. 방어가 아닌 투기."),
    ("결론 6", "진짜 방어주는 보험 섹터", "5일 내내 초과수익률 양수. Q1비율 42.9%. 절대변동성 낮고 거래대금 충분."),
]
for num, title, desc in conclusions:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell = ws.cell(row=r, column=1, value=f"{num}: {title}")
    cell.font = Font("Arial", bold=True, size=10, color="f59e0b"); cell.fill = BG0
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    cell = ws.cell(row=r, column=1, value=desc)
    cell.font = NRM; cell.fill = BG1; cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 30; r += 2

# ── Sheet 2: 시총 5분위 ──
ws2 = wb.create_sheet("시총5분위_핵심"); ws2.sheet_properties.tabColor = "3b82f6"
r = 1
for m in ["KOSPI", "KOSDAQ"]:
    r = stitle(ws2, r, f"{m} — 시총 5분위: 하락avg vs 반등(3/5) 동등비교", 10, "3b82f6"); r += 1
    hdr(ws2, r, ["시총그룹", "종목수", "시총범위(억)", "하락avg(%)", "반등3/5(%)", "순합(%)", "횡보3/6(%)", "하락양수%", "반등양수%", "Q1비율%"],
        [10, 8, 20, 12, 12, 10, 10, 10, 10, 10]); r += 1
    sub = meta[meta["시장구분"] == m]
    for g in ["1_최소", "2_소형", "3_중형", "4_대형", "5_최대"]:
        gs = sub[sub["Q5"] == g]
        lo, hi = gs["시총_억"].min(), gs["시총_억"].max()
        d = gs["하락avg"].mean(); rv = gs["반등"].mean(); h = gs["횡보"].mean()
        dp = (gs["하락avg"] > 0).mean() * 100; rp = (gs["반등"] > 0).mean() * 100
        q1 = (gs["사분면"] == "Q1").mean() * 100
        wrow(ws2, r, [g, len(gs), f"{lo:,.0f}~{hi:,.0f}", d, rv, d+rv, h, dp, rp, q1], pn=[4, 5, 6, 7]); r += 1
    r += 1

    r = stitle(ws2, r, f"{m} — 시총 5분위: 반등매매 (3/4매수→3/5매도)", 10, "10b981"); r += 1
    hdr(ws2, r, ["시총그룹", "종목수", "3/4하락(%)", "1일반등(%)", "2일반등(%)", "1일승률(%)", "2일승률(%)", "회복률(%)", "시총중앙(억)"],
        [10, 8, 12, 12, 12, 10, 10, 10, 12]); r += 1
    for g in ["1_최소", "2_소형", "3_중형", "4_대형", "5_최대"]:
        gs = sub[sub["Q5"] == g]
        d04 = gs["rt_04"].mean(); b1 = gs["반등1일"].mean(); b2 = gs["반등2일"].mean()
        wr1 = (gs["반등1일"] > 0).mean() * 100; wr2 = (gs["반등2일"] > 0).mean() * 100
        rec = (b1 / abs(d04) * 100) if d04 != 0 else 0
        wrow(ws2, r, [g, len(gs), d04, b1, b2, wr1, wr2, rec, gs["시총_억"].median()], pn=[3, 4, 5]); r += 1
    r += 1

    r = stitle(ws2, r, f"{m} — 시총 5분위: 하락 취약도 (등락률 기준)", 10, "ef4444"); r += 1
    hdr(ws2, r, ["시총그룹", "종목수", "3/3(%)", "3/4(%)", "3/9(%)", "3일합산(%)", "3/5반등(%)", "시총중앙(억)"],
        [10, 8, 10, 10, 10, 12, 12, 12]); r += 1
    for g in ["1_최소", "2_소형", "3_중형", "4_대형", "5_최대"]:
        gs = sub[sub["Q5"] == g]
        d03 = gs["rt_03"].mean(); d04 = gs["rt_04"].mean(); d09 = gs["rt_09"].mean(); d05 = gs["rt_05"].mean()
        wrow(ws2, r, [g, len(gs), d03, d04, d09, d03+d04+d09, d05, gs["시총_억"].median()], pn=[3, 4, 5, 6, 7]); r += 1
    r += 2

# ── Sheet 3: 유동성 착시 ──
ws3 = wb.create_sheet("유동성착시검증"); ws3.sheet_properties.tabColor = "f59e0b"
r = 1
for m in ["KOSPI", "KOSDAQ"]:
    r = stitle(ws3, r, f"{m} — 거래대금 5분위별 초과수익률 양수 비율 (착시 판별)", 12, "f59e0b"); r += 1
    hdr(ws3, r, ["유동성", "종목수", "거래대금중앙(억)", "시총중앙(억)", "하락avg(%)", "반등3/5(%)", "순합(%)",
                  "하락양수%", "반등양수%", "3/3양수%", "3/4양수%", "3/5양수%"],
        [10, 8, 14, 12, 12, 12, 10, 10, 10, 10, 10, 10]); r += 1
    sub = meta[meta["시장구분"] == m]
    for g in ["1_극소", "2_소", "3_중", "4_대", "5_극대"]:
        gs = sub[sub["유동성5"] == g]
        d = gs["하락avg"].mean(); rv = gs["반등"].mean()
        dp = (gs["하락avg"] > 0).mean() * 100; rp = (gs["반등"] > 0).mean() * 100
        a03 = (gs["ex_03"] > 0).mean() * 100; a04 = (gs["ex_04"] > 0).mean() * 100; a05 = (gs["ex_05"] > 0).mean() * 100
        wrow(ws3, r, [g, len(gs), gs["평균거래대금_억"].median(), gs["시총_억"].median(), d, rv, d+rv, dp, rp, a03, a04, a05], pn=[5, 6, 7]); r += 1
    r += 2

# ── Sheet 4: 섹터별 ──
ws4 = wb.create_sheet("섹터분석"); ws4.sheet_properties.tabColor = "8b5cf6"
r = 1
for m in ["KOSPI", "KOSDAQ"]:
    r = stitle(ws4, r, f"{m} — 섹터별 동등비교 (하락avg + 반등3/5)", 12, "8b5cf6"); r += 1
    hdr(ws4, r, ["업종", "종목수", "하락avg(%)", "반등3/5(%)", "순합(%)", "횡보3/6(%)", "Q1비율%", "시총중앙(억)",
                  "3/3(%)", "3/4(%)", "3/5(%)", "3/9(%)"],
        [16, 8, 12, 12, 10, 10, 8, 12, 8, 8, 8, 8]); r += 1
    sub = meta[meta["시장구분"] == m]
    sec = sub.groupby("업종명").agg(
        n=("종목코드", "count"), d=("하락avg", "mean"), rv=("반등", "mean"), h=("횡보", "mean"),
        q1=("사분면", lambda x: (x == "Q1").mean() * 100), cap=("시총_억", "median"),
        e03=("ex_03", "mean"), e04=("ex_04", "mean"), e05=("ex_05", "mean"), e09=("ex_09", "mean"),
    ).round(2)
    sec["순합"] = sec["d"] + sec["rv"]
    sec = sec.sort_values("순합", ascending=False)
    for s, row in sec.iterrows():
        wrow(ws4, r, [s, row["n"], row["d"], row["rv"], row["순합"], row["h"], row["q1"], row["cap"],
                      row["e03"], row["e04"], row["e05"], row["e09"]], pn=[3, 4, 5, 6, 9, 10, 11, 12]); r += 1
    r += 2

# ── Sheet 5: 이상패턴 ──
ws5 = wb.create_sheet("이상패턴종목"); ws5.sheet_properties.tabColor = "10b981"
r = 1
cols = ["종목명", "시장", "업종", "시총(억)", "거래대금(억)", "ex3/3", "ex3/4", "ex3/5", "ex3/6", "ex3/9", "하락avg", "반등3/5", "동등스코어", "사분면"]
widths = [14, 8, 14, 10, 10, 8, 8, 8, 8, 8, 10, 10, 10, 6]
pn = [6, 7, 8, 9, 10, 11, 12, 13]

rebels = meta[(meta["ex_03"] > 5) & (meta["ex_04"] > 5) & (meta["ex_09"] > 5)].sort_values("하락avg", ascending=False)
r = stitle(ws5, r, f"역행상승: 폭락 3일 모두 초과 >5% ({len(rebels)}종목)", 14, "10b981"); r += 1
hdr(ws5, r, cols, widths); r += 1
for _, row in rebels.iterrows():
    wrow(ws5, r, [row["종목명"], row["시장구분"], row["업종명"], row["시총_억"], row["평균거래대금_억"],
                  row["ex_03"], row["ex_04"], row["ex_05"], row["ex_06"], row["ex_09"],
                  row["하락avg"], row["반등"], row["동등스코어"], row["사분면"]], pn=pn); r += 1

r += 2
anti = meta[(meta["ex_05"] < -5) & (meta["ex_06"] < -5)].sort_values("반등")
r = stitle(ws5, r, f"반등역행: 반등+횡보 둘다 초과 <-5% ({len(anti)}종목)", 14, "ef4444"); r += 1
hdr(ws5, r, cols, widths); r += 1
for _, row in anti.iterrows():
    wrow(ws5, r, [row["종목명"], row["시장구분"], row["업종명"], row["시총_억"], row["평균거래대금_억"],
                  row["ex_03"], row["ex_04"], row["ex_05"], row["ex_06"], row["ex_09"],
                  row["하락avg"], row["반등"], row["동등스코어"], row["사분면"]], pn=pn); r += 1

# ── Sheet 6: 전체 데이터 ──
ws6 = wb.create_sheet("전체종목"); ws6.sheet_properties.tabColor = "94a3b8"
r = 1
fc = ["종목코드", "종목명", "시장", "업종", "시총그룹", "시총(억)", "거래대금(억)", "절대변동성",
      "ex3/3", "ex3/4", "ex3/5", "ex3/6", "ex3/9", "rt3/3", "rt3/4", "rt3/5", "rt3/6", "rt3/9",
      "하락avg", "반등3/5", "동등스코어", "사분면", "반등1일", "반등2일"]
fw = [10, 14, 8, 14, 8, 10, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 10, 10, 10, 6, 10, 10]
hdr(ws6, r, fc, fw); r += 1
for _, row in meta.sort_values("동등스코어", ascending=False).iterrows():
    wrow(ws6, r, [
        row["종목코드"], row["종목명"], row["시장구분"], row["업종명"], str(row.get("Q5", "")),
        row["시총_억"], row["평균거래대금_억"], row["절대변동성"],
        row.get("ex_03"), row.get("ex_04"), row.get("ex_05"), row.get("ex_06"), row.get("ex_09"),
        row.get("rt_03"), row.get("rt_04"), row.get("rt_05"), row.get("rt_06"), row.get("rt_09"),
        row["하락avg"], row["반등"], row["동등스코어"], row["사분면"],
        row.get("반등1일"), row.get("반등2일"),
    ], pn=[9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 23, 24]); r += 1

if "Sheet" in wb.sheetnames: del wb["Sheet"]
for n in wb.sheetnames: wb[n].freeze_panes = "A3"
wb.save(EXCEL_OUT)
print(f"✅ Excel: {EXCEL_OUT}")

# ══════════════════════════════════════
# HTML
# ══════════════════════════════════════
def grp_data(market):
    sub = meta[meta["시장구분"] == market]
    rows = []
    for g in ["1_최소", "2_소형", "3_중형", "4_대형", "5_최대"]:
        gs = sub[sub["Q5"] == g]
        lo, hi = gs["시총_억"].min(), gs["시총_억"].max()
        d = gs["하락avg"].mean(); rv = gs["반등"].mean()
        d04 = gs["rt_04"].mean(); b1 = gs["반등1일"].mean()
        wr1 = (gs["반등1일"] > 0).mean() * 100
        dp = (gs["하락avg"] > 0).mean() * 100; rp = (gs["반등"] > 0).mean() * 100
        rows.append({"g": g.split("_")[1], "n": len(gs), "range": f"{lo:,.0f}~{hi:,.0f}",
            "d": round(d, 2), "r": round(rv, 2), "net": round(d+rv, 2),
            "d04": round(d04, 1), "b1": round(b1, 1), "wr": round(wr1, 1), "dp": round(dp, 1), "rp": round(rp, 1)})
    return rows

kospi_q5 = grp_data("KOSPI"); kosdaq_q5 = grp_data("KOSDAQ")

def sec_data(market, n=8):
    sub = meta[meta["시장구분"] == market]
    sec = sub.groupby("업종명").agg(n=("종목코드", "count"), d=("하락avg", "mean"), r=("반등", "mean"),
        q1=("사분면", lambda x: (x == "Q1").mean() * 100)).round(2)
    sec["net"] = (sec["d"] + sec["r"]).round(2)
    return sec.sort_values("net", ascending=False).head(n).reset_index().to_dict("records")

ks_top = sec_data("KOSPI")

def liq_data(market):
    sub = meta[meta["시장구분"] == market]
    rows = []
    for g in ["1_극소", "2_소", "3_중", "4_대", "5_극대"]:
        gs = sub[sub["유동성5"] == g]
        dp = (gs["하락avg"] > 0).mean() * 100; rp = (gs["반등"] > 0).mean() * 100
        rows.append({"g": g.split("_")[1], "n": len(gs), "dp": round(dp, 1), "rp": round(rp, 1)})
    return rows

ks_liq = liq_data("KOSPI"); kq_liq = liq_data("KOSDAQ")

def trow(r, cls=""):
    return f'<tr><td>{r["g"]}</td><td>{r["n"]}</td><td class="{cls if r.get("dp",0)>70 else ""}">{r.get("dp","")}{"%" if "dp" in r else ""}</td><td class="{"neg" if r.get("rp",0)<20 else ""}">{r.get("rp","")}{"%" if "rp" in r else ""}</td></tr>'

html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>미·이란 전쟁 위기 한국 증시 분석</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#070d1b;color:#e2e8f0;font-family:'Noto Sans KR',sans-serif;line-height:1.6}}
.slide{{min-height:100vh;padding:48px 60px;display:flex;flex-direction:column;justify-content:center;border-bottom:1px solid #1e293b}}
.slide:nth-child(even){{background:#0b1120}}
h1{{font-size:42px;font-weight:900;letter-spacing:-1.5px;margin-bottom:12px}}
h1 span{{color:#ef4444}}
h2{{font-size:28px;font-weight:700;margin-bottom:24px;letter-spacing:-0.5px}}
h2 .num{{display:inline-block;background:#3b82f6;color:#fff;width:36px;height:36px;border-radius:50%;text-align:center;line-height:36px;font-size:16px;margin-right:10px}}
.sub{{color:#64748b;font-size:14px;margin-bottom:32px}}
.hl{{background:linear-gradient(120deg,rgba(239,68,68,0.15),rgba(59,130,246,0.15));border-left:4px solid #ef4444;padding:16px 24px;border-radius:0 8px 8px 0;margin:20px 0;font-size:15px}}
.hl b{{color:#f8fafc}}
.sr{{display:flex;gap:12px;flex-wrap:wrap;margin:16px 0}}
.st{{background:#1e293b;border-radius:10px;padding:16px 20px;flex:1;min-width:140px;border:1px solid #334155}}
.st .lb{{font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:0.5px}}
.st .vl{{font-size:28px;font-weight:700;margin:4px 0}}
.st .ds{{font-size:11px;color:#94a3b8}}
.pos{{color:#10b981}}.neg{{color:#ef4444}}.warn{{color:#f59e0b}}
table{{width:100%;border-collapse:collapse;margin:12px 0;font-size:13px}}
th{{background:#1e293b;color:#94a3b8;padding:10px 12px;text-align:center;font-weight:600;border-bottom:2px solid #334155}}
td{{padding:8px 12px;text-align:center;border-bottom:1px solid #1e293b}}
tr:hover{{background:rgba(59,130,246,0.05)}}
.tl{{text-align:left}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:24px}}
.cd{{background:rgba(30,41,59,0.5);border-radius:12px;padding:20px;border:1px solid #1e293b}}
.ct{{font-size:14px;font-weight:700;margin-bottom:12px}}
@media print{{.slide{{page-break-after:always;min-height:auto;padding:24px}}}}
</style>
</head>
<body>

<div class="slide" style="justify-content:center;align-items:center;text-align:center">
<h1>🇮🇷 미·이란 전쟁 위기<br><span>한국 증시</span> 심층 분석</h1>
<p class="sub" style="font-size:16px">2026.03.03~03.09 · 5거래일 · 2,771종목<br>KOSPI→KOSPI지수 / KOSDAQ→KOSDAQ지수 대비 초과수익률</p>
<div class="sr" style="justify-content:center">
<div class="st" style="max-width:140px"><div class="lb">분석 기간</div><div class="vl" style="font-size:20px">5일</div><div class="ds">하락3 반등1 횡보1</div></div>
<div class="st" style="max-width:140px"><div class="lb">종목수</div><div class="vl" style="font-size:20px">2,771</div><div class="ds">KOSPI 951 + KOSDAQ 1,820</div></div>
<div class="st" style="max-width:140px"><div class="lb">최대 하락</div><div class="vl neg" style="font-size:20px">-14.0%</div><div class="ds">KOSDAQ 3/4</div></div>
<div class="st" style="max-width:140px"><div class="lb">최대 반등</div><div class="vl pos" style="font-size:20px">+14.1%</div><div class="ds">KOSDAQ 3/5</div></div>
</div></div>

<div class="slide">
<h2>시장 개요</h2>
<div class="sr">
<div class="st"><div class="lb">3/3(월) 폭락1</div><div class="vl neg">-7.24%</div><div class="ds">KOSDAQ -4.62%</div></div>
<div class="st"><div class="lb">3/4(화) 폭락2</div><div class="vl neg">-12.06%</div><div class="ds">KOSDAQ -14.00%</div></div>
<div class="st"><div class="lb">3/5(수) 반등</div><div class="vl pos">+9.63%</div><div class="ds">KOSDAQ +14.10%</div></div>
<div class="st"><div class="lb">3/6(목) 횡보</div><div class="vl" style="color:#94a3b8">+0.02%</div><div class="ds">KOSDAQ +3.43%</div></div>
<div class="st"><div class="lb">3/9(월) 재폭락</div><div class="vl neg">-5.96%</div><div class="ds">KOSDAQ -4.54%</div></div>
</div>
<div class="hl"><b>하락 3일 + 반등 1일 + 횡보 1일.</b> 3/6은 KOSPI +0.02%로 실질 횡보.<br>→ <b>하락(3/3,3/4,3/9) 3일 평균 vs 반등(3/5만)</b>으로 동등비교.</div>
</div>

<div class="slide">
<h2><span class="num">1</span>소형주 방어력은 <span class="neg">착시</span></h2>
<div class="g2">
<div class="cd"><div class="ct">KOSPI 거래대금별 초과수익률 양수 비율</div>
<table><tr><th>유동성</th><th>종목</th><th>하락일 양수%</th><th>반등일 양수%</th></tr>
{"".join(f'<tr><td>{r["g"]}</td><td>{r["n"]}</td><td class="{"pos" if r["dp"]>70 else ""}">{r["dp"]}%</td><td class="{"neg" if r["rp"]<20 else ""}">{r["rp"]}%</td></tr>' for r in ks_liq)}</table></div>
<div class="cd"><div class="ct">KOSDAQ 거래대금별 초과수익률 양수 비율</div>
<table><tr><th>유동성</th><th>종목</th><th>하락일 양수%</th><th>반등일 양수%</th></tr>
{"".join(f'<tr><td>{r["g"]}</td><td>{r["n"]}</td><td class="{"pos" if r["dp"]>70 else ""}">{r["dp"]}%</td><td class="{"neg" if r["rp"]<20 else ""}">{r["rp"]}%</td></tr>' for r in kq_liq)}</table></div>
</div>
<div class="hl">KOSDAQ 극소유동: 하락일 양수 <b class="pos">98.1%</b> → 반등일 양수 <b class="neg">0.0%</b><br>거래가 없어서 가격이 안 움직인 것. 초과수익률만 보면 '방어력'으로 오해.</div>
</div>

<div class="slide">
<h2><span class="num">2</span>동등비교하면 시총별 <span class="warn">알파 없음</span></h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI 시총 5분위</div>
<table><tr><th>그룹</th><th>종목</th><th>하락avg</th><th>반등3/5</th><th>순합</th></tr>
{"".join(f'<tr><td>{r["g"]}</td><td>{r["n"]}</td><td class="{"pos" if r["d"]>0 else "neg"}">{r["d"]:+.2f}%</td><td class="{"pos" if r["r"]>0 else "neg"}">{r["r"]:+.2f}%</td><td class="{"pos" if r["net"]>0 else "neg"}"><b>{r["net"]:+.2f}%</b></td></tr>' for r in kospi_q5)}</table></div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ 시총 5분위</div>
<table><tr><th>그룹</th><th>종목</th><th>하락avg</th><th>반등3/5</th><th>순합</th></tr>
{"".join(f'<tr><td>{r["g"]}</td><td>{r["n"]}</td><td class="{"pos" if r["d"]>0 else "neg"}">{r["d"]:+.2f}%</td><td class="{"pos" if r["r"]>0 else "neg"}">{r["r"]:+.2f}%</td><td class="{"pos" if r["net"]>0 else "neg"}"><b>{r["net"]:+.2f}%</b></td></tr>' for r in kosdaq_q5)}</table></div>
</div>
<div class="hl">KOSPI: 전 그룹 순합 <b>±0.25% 이내</b>. 하락일에 덜 빠진 만큼 반등일에 정확히 덜 오름. 상관계수 <b>-0.989</b>.<br>KOSDAQ: 전 그룹 <b class="neg">마이너스</b>. 소형주가 가장 큰 손실 (-5.82%).</div>
</div>

<div class="slide">
<h2><span class="num">3</span>반등매매는 <span class="pos">대형주</span>가 최적</h2>
<div class="g2">
<div class="cd"><div class="ct" style="color:#3b82f6">KOSPI: 3/4매수→3/5매도</div>
<table><tr><th>그룹</th><th>3/4하락</th><th>반등</th><th>승률</th></tr>
{"".join(f'<tr><td>{r["g"]}</td><td class="neg">{r["d04"]:+.1f}%</td><td class="pos">+{r["b1"]:.1f}%</td><td>{r["wr"]:.1f}%</td></tr>' for r in kospi_q5)}</table></div>
<div class="cd"><div class="ct" style="color:#f97316">KOSDAQ: 3/4매수→3/5매도</div>
<table><tr><th>그룹</th><th>3/4하락</th><th>반등</th><th>승률</th></tr>
{"".join(f'<tr><td>{r["g"]}</td><td class="neg">{r["d04"]:+.1f}%</td><td class="pos">+{r["b1"]:.1f}%</td><td>{r["wr"]:.1f}%</td></tr>' for r in kosdaq_q5)}</table></div>
</div>
<div class="hl">시총 클수록 반등 수익·승률 <b>단조증가</b> (예외 없음).<br>KOSDAQ 최대: <b class="neg">-14.0%</b> → <b class="pos">+13.2%</b>, 승률 <b>99.5%</b>.</div>
</div>

<div class="slide">
<h2><span class="num">4</span>섹터: 보험이 유일한 <span class="pos">진짜 방어주</span></h2>
<div class="g2">
<div class="cd"><div class="ct">KOSPI 섹터 순합 TOP</div>
<table><tr><th class="tl">업종</th><th>종목</th><th>하락avg</th><th>반등3/5</th><th>순합</th><th>Q1%</th></tr>
{"".join(f'<tr><td class="tl">{r["업종명"]}</td><td>{r["n"]}</td><td class="{"pos" if r["d"]>0 else "neg"}">{r["d"]:+.2f}%</td><td class="{"pos" if r["r"]>0 else "neg"}">{r["r"]:+.2f}%</td><td class="{"pos" if r["net"]>0 else "neg"}"><b>{r["net"]:+.2f}%</b></td><td>{r["q1"]:.0f}%</td></tr>' for r in ks_top)}</table></div>
<div class="cd"><div class="ct">에너지/석유 테마의 함정</div>
<p style="font-size:13px;color:#94a3b8;margin-bottom:12px">하락일 +30% 초과수익 → 반등일 -10~20% 되돌림</p>
<table><tr><th class="tl">종목</th><th>하락avg</th><th>반등3/5</th><th>순합</th></tr>
<tr><td class="tl">대성에너지</td><td class="pos">+35.5%</td><td class="neg">-16.1%</td><td class="pos">+19.4%</td></tr>
<tr><td class="tl">극동유화</td><td class="pos">+30.8%</td><td class="neg">-27.7%</td><td class="pos">+3.1%</td></tr>
<tr><td class="tl">흥구석유</td><td class="pos">+28.1%</td><td class="neg">-18.3%</td><td class="pos">+9.7%</td></tr></table>
<p style="font-size:12px;color:#f59e0b;margin-top:8px">⚠ 방어가 아닌 역방향 투기. 시장 안정 시 급락.</p></div>
</div>
<div class="hl"><b style="color:#10b981">보험 섹터</b>: 5일 내내 초과수익률 양수 (+3.1, +2.2, -0.5, +2.0, +2.8). Q1비율 42.9%로 전 섹터 1위.</div>
</div>

<div class="slide">
<h2><span class="num">5</span>두 폭락은 <span class="warn">성격이 달랐다</span></h2>
<div class="g2">
<div class="cd"><div class="ct"><span style="background:rgba(239,68,68,0.15);color:#f87171;padding:2px 8px;border-radius:4px;font-size:11px">3/4 폭락</span> 무차별 투매</div>
<ul style="font-size:13px;list-style:none;padding:0">
<li style="margin:8px 0">• KOSPI <b class="neg">-12.06%</b>, KOSDAQ <b class="neg">-14.00%</b></li>
<li style="margin:8px 0">• 전 섹터 거래대금 급증 (비금속 +404%, 에너지 +413%)</li>
<li style="margin:8px 0">• 소형주까지 전부 밀림 — 패닉</li></ul></div>
<div class="cd"><div class="ct"><span style="background:rgba(245,158,11,0.15);color:#fbbf24;padding:2px 8px;border-radius:4px;font-size:11px">3/9 폭락</span> 선별적 매도</div>
<ul style="font-size:13px;list-style:none;padding:0">
<li style="margin:8px 0">• KOSPI <b class="neg">-5.96%</b>, KOSDAQ <b class="neg">-4.54%</b></li>
<li style="margin:8px 0">• 3/4에 버틴 종목이 3/9에 무너짐 (스튜디오산타: +14%→-67%)</li>
<li style="margin:8px 0">• 반대로 3/4에 무너진 종목이 3/9에 방어 (SK가스, 남해화학)</li></ul></div>
</div>
<div class="hl">3/4: "다 팔아!" → 3/9: "골라서 팔자." <b>시장이 학습함.</b></div>
</div>

<div class="slide" style="align-items:center;text-align:center">
<h2 style="font-size:32px">최종 결론</h2>
<div style="max-width:700px;text-align:left;margin:24px auto">
<div style="display:flex;gap:16px;align-items:flex-start;margin:20px 0;padding:16px;background:#1e293b;border-radius:10px">
<div style="font-size:24px">📉</div>
<div><b style="font-size:15px">폭락 반등매매를 할 거면</b><br><span style="color:#94a3b8;font-size:13px">시총이 크면 클수록 좋다. KOSDAQ 최대 시총 그룹이 수익률·승률·회복률 모두 최적. 1일 트레이딩만 유효.</span></div></div>
<div style="display:flex;gap:16px;align-items:flex-start;margin:20px 0;padding:16px;background:#1e293b;border-radius:10px">
<div style="font-size:24px">🛡️</div>
<div><b style="font-size:15px">방어 포트폴리오를 구성할 거면</b><br><span style="color:#94a3b8;font-size:13px">보험 섹터가 유일한 진짜 방어주. 시총으로는 전체기간 알파를 만들 수 없음 (동등비교 시 전 그룹 ±0%).</span></div></div>
<div style="display:flex;gap:16px;align-items:flex-start;margin:20px 0;padding:16px;background:#1e293b;border-radius:10px">
<div style="font-size:24px">⚠️</div>
<div><b style="font-size:15px">주의할 것</b><br><span style="color:#94a3b8;font-size:13px">소형주·저유동 종목의 '방어력'은 착시. 에너지 테마는 역방향 베팅이지 방어가 아님. 초과수익률은 반드시 거래대금과 같이 볼 것.</span></div></div>
</div></div>

</body></html>"""

with open(HTML_OUT, "w", encoding="utf-8") as f:
    f.write(html)
print(f"✅ HTML: {HTML_OUT}")
print("Done!")