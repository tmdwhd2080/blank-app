# -*- coding: utf-8 -*-
"""
FX Risk Regime 분석 결과 시각화 모듈
- 엑셀 리포트 생성
- 그래프 시각화

레짐 해석 (수정됨!):
    STATES = 1  → Risk-On  → 안정기
    STATES = 0  → Neutral  → 평범
    STATES = -1 → Risk-Off → 위기
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from pathlib import Path

# 한글 폰트 설정
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

# openpyxl 사용
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference, LineChart
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[경고] openpyxl 없음. pip install openpyxl 실행하세요.")


# 레짐 매핑
REGIME_MAP = {
    1: '안정(Risk-On)',
    0: '평범(Neutral)',
    -1: '위기(Risk-Off)'
}


# ============================================================
# 1. 그래프 시각화
# ============================================================

def plot_regime_beta_distribution(df_results, save_path=None):
    """
    레짐별 β_fx 분포 박스플롯
    """
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # 순서: 안정 → 평범 → 위기
    data = [
        df_results['beta_fx_stable'].dropna(),
        df_results['beta_fx_normal'].dropna(),
        df_results['beta_fx_crisis'].dropna()
    ]
    
    bp = ax.boxplot(data, labels=['안정 (Risk-On)', '평범 (Neutral)', '위기 (Risk-Off)'], patch_artist=True)
    
    colors = ['#6bcb77', '#ffd93d', '#ff6b6b']  # 초록(안정), 노랑(평범), 빨강(위기)
    for patch, color in zip(bp['boxes'], colors):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)
    
    ax.axhline(y=0, color='black', linestyle='--', alpha=0.5)
    ax.set_ylabel('β_fx (환율 민감도)')
    ax.set_xlabel('레짐')
    ax.set_title('레짐별 FX 민감도 분포 (대칭)')
    ax.grid(axis='y', alpha=0.3)
    
    # 평균값 표시
    means = [d.mean() for d in data]
    for i, mean in enumerate(means):
        ax.scatter(i+1, mean, color='red', s=100, zorder=5, marker='D')
        ax.annotate(f'{mean:.3f}', (i+1, mean), textcoords="offset points", 
                   xytext=(10, 5), fontsize=10, color='red')
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"    저장: {save_path}")
    
    plt.close()
    return fig


def plot_asymmetric_beta_comparison(df_results_asym, save_path=None):
    """
    비대칭 β_fx 비교 (상승 vs 하락)
    """
    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    
    # 순서: 안정 → 평범 → 위기
    regimes = [('stable', '안정(Risk-On)'), ('normal', '평범'), ('crisis', '위기(Risk-Off)')]
    colors = ['#6bcb77', '#ffd93d', '#ff6b6b']
    
    for ax, (regime, name), color in zip(axes, regimes, colors):
        up_col = f'beta_fx_{regime}_up'
        down_col = f'beta_fx_{regime}_down'
        
        up_data = df_results_asym[up_col].dropna()
        down_data = df_results_asym[down_col].dropna()
        
        bp = ax.boxplot([up_data, down_data], labels=['환율↑', '환율↓'], patch_artist=True)
        
        bp['boxes'][0].set_facecolor('#ff9999')
        bp['boxes'][1].set_facecolor('#99ccff')
        
        ax.axhline(y=0, color='black', linestyle='--', alpha=0.5)
        ax.set_ylabel('β_fx')
        ax.set_title(f'{name}')
        ax.grid(axis='y', alpha=0.3)
        
        # 평균값 표시
        for i, (data, label) in enumerate([(up_data, '상승'), (down_data, '하락')]):
            mean = data.mean()
            ax.scatter(i+1, mean, color='red', s=80, zorder=5, marker='D')
            ax.annotate(f'{mean:.3f}', (i+1, mean), textcoords="offset points",
                       xytext=(8, 3), fontsize=9, color='red')
    
    plt.suptitle('레짐별 비대칭 FX 민감도 (환율 상승 vs 하락)', fontsize=12, fontweight='bold')
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"    저장: {save_path}")
    
    plt.close()
    return fig


def plot_top_sensitive_stocks(df_results, regime='crisis', top_n=15, save_path=None):
    """
    FX 민감도 상위 종목 막대 그래프
    """
    col = f'beta_fx_{regime}'
    pval_col = f'pval_fx_{regime}'
    
    regime_names = {'stable': '안정(Risk-On)', 'normal': '평범', 'crisis': '위기(Risk-Off)'}
    regime_name = regime_names.get(regime, regime)
    
    # 유의한 종목만, 절대값 기준 상위
    df_sig = df_results[df_results[pval_col] < 0.1].copy()
    df_sig['abs_beta'] = df_sig[col].abs()
    df_top = df_sig.nlargest(top_n, 'abs_beta')
    
    if len(df_top) == 0:
        print(f"[{regime_name}] 유의한 종목 없음")
        return None
    
    fig, ax = plt.subplots(figsize=(12, 8))
    
    colors = ['#ff6b6b' if x < 0 else '#6bcb77' for x in df_top[col]]
    
    y_pos = range(len(df_top))
    ax.barh(y_pos, df_top[col], color=colors, alpha=0.8)
    
    ax.set_yticks(y_pos)
    ax.set_yticklabels(df_top['name'])
    ax.axvline(x=0, color='black', linewidth=0.8)
    ax.set_xlabel('β_fx (환율 민감도)')
    ax.set_title(f'{regime_name} 레짐 - FX 민감도 상위 종목 (Top {top_n})')
    ax.grid(axis='x', alpha=0.3)
    
    # 값 표시
    for i, (v, p) in enumerate(zip(df_top[col], df_top[pval_col])):
        sig = '***' if p < 0.01 else '**' if p < 0.05 else '*'
        offset = 0.02 if v >= 0 else -0.02
        ha = 'left' if v >= 0 else 'right'
        ax.text(v + offset, i, f'{v:.3f}{sig}', va='center', ha=ha, fontsize=9)
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"    저장: {save_path}")
    
    plt.close()
    return fig


def plot_regime_heatmap(df_results, save_path=None):
    """
    섹터별/레짐별 평균 β_fx 히트맵
    """
    fig, ax = plt.subplots(figsize=(10, 8))
    
    # 상위 30 종목 선택 (Wald test 유의한 종목 우선)
    df_sig = df_results[df_results['wald_p'] < 0.1].nsmallest(30, 'wald_p')
    
    if len(df_sig) < 5:
        df_sig = df_results.nsmallest(30, 'wald_p')
    
    # 히트맵 데이터 (순서: 안정 → 평범 → 위기)
    heatmap_data = df_sig[['name', 'beta_fx_stable', 'beta_fx_normal', 'beta_fx_crisis']].copy()
    heatmap_data = heatmap_data.set_index('name')
    heatmap_data.columns = ['안정', '평범', '위기']
    
    # 히트맵 그리기
    im = ax.imshow(heatmap_data.values, cmap='RdYlGn_r', aspect='auto')
    
    ax.set_xticks(range(3))
    ax.set_xticklabels(['안정(Risk-On)', '평범', '위기(Risk-Off)'])
    ax.set_yticks(range(len(heatmap_data)))
    ax.set_yticklabels(heatmap_data.index, fontsize=8)
    
    # 값 표시
    for i in range(len(heatmap_data)):
        for j in range(3):
            val = heatmap_data.iloc[i, j]
            color = 'white' if abs(val) > 0.5 else 'black'
            ax.text(j, i, f'{val:.2f}', ha='center', va='center', color=color, fontsize=8)
    
    ax.set_title('종목별 × 레짐별 FX 민감도 히트맵\n(Wald Test 유의 종목 상위 30)')
    plt.colorbar(im, ax=ax, label='β_fx')
    
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"    저장: {save_path}")
    
    plt.close()
    return fig


def plot_prediction_impact(df_pred, current_regime_info, fx_change, top_n=20, save_path=None):
    """
    예측 영향 그래프
    """
    regime_name = current_regime_info['regime_name']
    
    fig, axes = plt.subplots(1, 2, figsize=(14, 8))
    
    # 부정적 영향 Top N
    df_worst = df_pred.head(top_n)
    ax1 = axes[0]
    colors1 = ['#ff6b6b'] * len(df_worst)
    ax1.barh(range(len(df_worst)), df_worst['predicted_impact'] * 100, color=colors1, alpha=0.8)
    ax1.set_yticks(range(len(df_worst)))
    ax1.set_yticklabels(df_worst['name'])
    ax1.axvline(x=0, color='black', linewidth=0.8)
    ax1.set_xlabel('예상 수익률 영향 (%)')
    ax1.set_title(f'환율 {fx_change*100:+.1f}% 시 부정적 영향 종목')
    ax1.invert_yaxis()
    
    # 긍정적 영향 Top N
    df_best = df_pred.tail(top_n).iloc[::-1]
    ax2 = axes[1]
    colors2 = ['#6bcb77'] * len(df_best)
    ax2.barh(range(len(df_best)), df_best['predicted_impact'] * 100, color=colors2, alpha=0.8)
    ax2.set_yticks(range(len(df_best)))
    ax2.set_yticklabels(df_best['name'])
    ax2.axvline(x=0, color='black', linewidth=0.8)
    ax2.set_xlabel('예상 수익률 영향 (%)')
    ax2.set_title(f'환율 {fx_change*100:+.1f}% 시 긍정적 영향 종목')
    ax2.invert_yaxis()
    
    plt.suptitle(f'현재 레짐: {regime_name} | 환율 변화: {fx_change*100:+.1f}%', 
                 fontsize=12, fontweight='bold')
    plt.tight_layout()
    
    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"    저장: {save_path}")
    
    plt.close()
    return fig


# ============================================================
# 2. 엑셀 리포트 생성
# ============================================================

def create_excel_report(df_results, df_results_asym, current_regime_info, output_path):
    """
    종합 엑셀 리포트 생성
    """
    if not EXCEL_AVAILABLE:
        print("[오류] openpyxl이 설치되어 있지 않습니다.")
        return
    
    wb = Workbook()
    
    # 스타일 정의
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # ============================================================
    # Sheet 1: 요약
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = '요약'
    
    # 제목
    ws_summary['A1'] = 'FX Risk Regime 분석 결과'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:E1')
    
    # 레짐 해석 설명
    ws_summary['A3'] = '[레짐 해석]'
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['A4'] = 'STATES=1: 안정 (Risk-On, 위험자산 선호)'
    ws_summary['A5'] = 'STATES=0: 평범 (Neutral)'
    ws_summary['A6'] = 'STATES=-1: 위기 (Risk-Off, 안전자산 선호)'
    
    # 현재 레짐 정보
    ws_summary['A8'] = '[현재 레짐 정보]'
    ws_summary['A8'].font = Font(bold=True)
    ws_summary['A9'] = '날짜'
    ws_summary['B9'] = str(current_regime_info['date'])[:10]
    ws_summary['A10'] = '레짐'
    ws_summary['B10'] = f"{current_regime_info['regime_name']} ({current_regime_info['regime']})"
    
    # 대칭 분석 요약
    ws_summary['A12'] = '[대칭 분석 요약]'
    ws_summary['A12'].font = Font(bold=True)
    
    ws_summary['A13'] = '총 종목 수'
    ws_summary['B13'] = len(df_results)
    ws_summary['A14'] = '평균 R²(전체)'
    ws_summary['B14'] = f"{df_results['r_squared'].mean():.4f}"
    ws_summary['A15'] = '평균 R²(시장만)'
    ws_summary['B15'] = f"{df_results['r2_mkt_only'].mean():.4f}"
    ws_summary['A16'] = '평균 R²(FX 추가분)'
    avg_fx_partial = df_results['r2_fx_partial'].mean()
    ws_summary['B16'] = f"{avg_fx_partial:.4f}"
    ws_summary['A17'] = 'Wald 유의 종목'
    wald_sig = (df_results['wald_p'] < 0.05).sum()
    wald_pct = (df_results['wald_p'] < 0.05).mean() * 100
    ws_summary['B17'] = f"{wald_sig}개 ({wald_pct:.1f}%)"
    
    # 레짐별 평균 β
    ws_summary['A19'] = '[레짐별 평균 β_fx]'
    ws_summary['A19'].font = Font(bold=True)
    ws_summary['A20'] = '레짐'
    ws_summary['B20'] = '평균 β'
    ws_summary['C20'] = '유의(5%)'
    ws_summary['D20'] = '음수 종목'
    
    for col in ['A20', 'B20', 'C20', 'D20']:
        ws_summary[col].font = header_font
        ws_summary[col].fill = header_fill
    
    row = 21
    # 순서: 안정 → 평범 → 위기
    for regime, name in [('stable', '안정(Risk-On)'), ('normal', '평범'), ('crisis', '위기(Risk-Off)')]:
        ws_summary[f'A{row}'] = name
        ws_summary[f'B{row}'] = f"{df_results[f'beta_fx_{regime}'].mean():.4f}"
        ws_summary[f'C{row}'] = (df_results[f'pval_fx_{regime}'] < 0.05).sum()
        ws_summary[f'D{row}'] = (df_results[f'beta_fx_{regime}'] < 0).sum()
        row += 1
    
    # 결론
    ws_summary['A25'] = '[결론]'
    ws_summary['A25'].font = Font(bold=True)
    if wald_pct >= 30:
        conclusion = '✅ 레짐 분리 효과적 - 레짐별 β_fx 분리 사용 권장'
    elif wald_pct >= 10:
        conclusion = '⚠️ 일부 종목만 유효 - 선별적 사용 권장'
    else:
        conclusion = '❌ 레짐 효과 없음 - 단일 β_fx 사용 권장'
    ws_summary['A26'] = conclusion
    
    # 열 너비 조정
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    
    # ============================================================
    # Sheet 2: 대칭 분석 결과
    # ============================================================
    ws_sym = wb.create_sheet('대칭분석')
    
    # 주요 컬럼만 선택 (순서: 안정 → 평범 → 위기)
    cols_sym = ['code', 'name', 'n_obs', 'r_squared', 'r2_mkt_only', 'r2_fx_partial',
                'beta_fx_stable', 'beta_fx_normal', 'beta_fx_crisis',
                'pval_fx_stable', 'pval_fx_normal', 'pval_fx_crisis',
                'wald_p']
    df_sym_export = df_results[cols_sym].copy()
    df_sym_export = df_sym_export.sort_values('wald_p')
    
    # 헤더
    headers = ['종목코드', '종목명', 'N', 'R²_전체', 'R²_시장만', 'R²_FX추가분',
               'β_안정', 'β_평범', 'β_위기',
               'p_안정', 'p_평범', 'p_위기', 'Wald_p']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_sym.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터
    for row_idx, row_data in enumerate(df_sym_export.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_sym.cell(row=row_idx, column=col_idx)
            if isinstance(value, float):
                cell.value = round(value, 4)
            else:
                cell.value = value
            
            # Wald p < 0.05면 강조
            if col_idx == 13 and isinstance(value, float) and value < 0.05:
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # 열 너비
    col_widths = [12, 15, 6, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws_sym.column_dimensions[chr(64+i)].width = width
    
    # ============================================================
    # Sheet 3: 비대칭 분석 결과
    # ============================================================
    ws_asym = wb.create_sheet('비대칭분석')
    
    cols_asym = ['code', 'name', 'n_obs', 'r_squared',
                 'beta_fx_stable_up', 'beta_fx_stable_down',
                 'beta_fx_normal_up', 'beta_fx_normal_down',
                 'beta_fx_crisis_up', 'beta_fx_crisis_down',
                 'wald_stable_asym_p', 'wald_normal_asym_p', 'wald_crisis_asym_p']
    
    df_asym_export = df_results_asym[[c for c in cols_asym if c in df_results_asym.columns]].copy()
    
    headers_asym = ['종목코드', '종목명', 'N', 'R²',
                    'β_안정↑', 'β_안정↓', 'β_평범↑', 'β_평범↓', 'β_위기↑', 'β_위기↓',
                    'p_안정비대칭', 'p_평범비대칭', 'p_위기비대칭']
    
    for col_idx, header in enumerate(headers_asym[:len(df_asym_export.columns)], 1):
        cell = ws_asym.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_idx, row_data in enumerate(df_asym_export.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_asym.cell(row=row_idx, column=col_idx)
            if isinstance(value, float):
                cell.value = round(value, 4)
            else:
                cell.value = value
    
    # ============================================================
    # Sheet 4: 레짐별 Top 종목
    # ============================================================
    ws_top = wb.create_sheet('Top종목')
    
    row = 1
    # 순서: 안정 → 평범 → 위기
    for regime, name in [('stable', '안정(Risk-On)'), ('normal', '평범'), ('crisis', '위기(Risk-Off)')]:
        ws_top.cell(row=row, column=1, value=f'{name} 레짐 - FX 민감도 Top 20')
        ws_top.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1
        
        # 헤더
        for col_idx, header in enumerate(['종목코드', '종목명', 'β_fx', 'p-value'], 1):
            cell = ws_top.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        
        # 데이터
        col = f'beta_fx_{regime}'
        pval_col = f'pval_fx_{regime}'
        df_top = df_results[df_results[pval_col] < 0.1].nsmallest(20, col)
        
        for _, stock in df_top.iterrows():
            ws_top.cell(row=row, column=1, value=stock['code'])
            ws_top.cell(row=row, column=2, value=stock['name'])
            ws_top.cell(row=row, column=3, value=round(stock[col], 4))
            ws_top.cell(row=row, column=4, value=round(stock[pval_col], 4))
            row += 1
        
        row += 2
    
    # 저장
    wb.save(output_path)
    print(f"    엑셀 저장: {output_path}")


# ============================================================
# 3. 전체 시각화 실행
# ============================================================

def generate_all_visualizations(df_results, df_results_asym, df_regime, results_dir):
    """
    모든 시각화 생성
    """
    from analysis import get_current_regime, predict_fx_impact
    
    results_dir = Path(results_dir)
    results_dir.mkdir(exist_ok=True)
    
    print("\n[그래프 생성]")
    
    # 1. 레짐별 β 분포
    plot_regime_beta_distribution(
        df_results, 
        save_path=results_dir / 'fig1_regime_beta_distribution.png'
    )
    
    # 2. 비대칭 비교
    plot_asymmetric_beta_comparison(
        df_results_asym,
        save_path=results_dir / 'fig2_asymmetric_comparison.png'
    )
    
    # 3. 레짐별 Top 종목 (순서: 안정 → 평범 → 위기)
    for regime in ['stable', 'normal', 'crisis']:
        plot_top_sensitive_stocks(
            df_results, 
            regime=regime,
            save_path=results_dir / f'fig3_top_stocks_{regime}.png'
        )
    
    # 4. 히트맵
    plot_regime_heatmap(
        df_results,
        save_path=results_dir / 'fig4_heatmap.png'
    )
    
    # 5. 예측 영향
    current = get_current_regime(df_regime)
    df_pred = predict_fx_impact(df_results, current['regime'], fx_change=0.01)
    plot_prediction_impact(
        df_pred, current, fx_change=0.01,
        save_path=results_dir / 'fig5_prediction_impact.png'
    )
    
    print("\n[엑셀 리포트 생성]")
    
    # 6. 엑셀 리포트
    create_excel_report(
        df_results, 
        df_results_asym,
        current,
        output_path=results_dir / 'fx_regime_report.xlsx'
    )
    
    print("\n모든 시각화 완료!")


if __name__ == '__main__':
    print("이 모듈은 main.py에서 import하여 사용합니다.")